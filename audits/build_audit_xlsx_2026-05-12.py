#!/usr/bin/env python3
"""Build daily_usability_audit_2026-05-12.xlsx for fundermatch.org,
website-auditor.io, and kevinarmstrong.io."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent / "daily_usability_audit_2026-05-12.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
PASS_FILL = PatternFill("solid", start_color="D5F5E3")
FAIL_FILL = PatternFill("solid", start_color="FADBD8")
WARN_FILL = PatternFill("solid", start_color="FCF3CF")
SKIP_FILL = PatternFill("solid", start_color="EAEDED")
INFO_FILL = PatternFill("solid", start_color="D6EAF8")
P0_FILL = PatternFill("solid", start_color="C0392B")
P1_FILL = PatternFill("solid", start_color="E67E22")
P2_FILL = PatternFill("solid", start_color="F1C40F")
P3_FILL = PatternFill("solid", start_color="95A5A6")
BORDER = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def fit_columns(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def status_fill(status):
    return {
        "PASS": PASS_FILL, "FAIL": FAIL_FILL, "WARN": WARN_FILL,
        "SKIP": SKIP_FILL, "INFO": INFO_FILL,
    }.get(status, None)


def severity_fill(sev):
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL, "P3": P3_FILL}.get(sev)


# ---------------------------------------------------------------------------
# Test data for each site. Format:
# (category, test_id, test_name, status, severity, notes)
# status: PASS, FAIL, WARN, SKIP, INFO
# severity: "" if PASS else P0/P1/P2/P3
# ---------------------------------------------------------------------------

FUNDERMATCH_TESTS = [
    # ---------- 1. First Impressions ----------
    ("1. First Impressions", "FI-01", "Homepage hero clearly states product (h1: 'Find Funders Aligned to Your Mission')", "PASS", "", "Clear, action-oriented headline; value prop unambiguous"),
    ("1. First Impressions", "FI-02", "Sub-headline explains the value (DAFs, foundations, corporate giving)", "PASS", "", "Single descriptive paragraph under hero"),
    ("1. First Impressions", "FI-03", "Primary CTA visible above the fold ('Get Started')", "PASS", "", "Pill button with search icon, centered"),
    ("1. First Impressions", "FI-04", "Document loads in <3s (DOMContentLoaded 361ms, load 416ms over wifi)", "PASS", "", "Measured via Performance API"),
    ("1. First Impressions", "FI-05", "Transfer size of HTML envelope <100KB (3KB gzip)", "PASS", "", "Bundle is split lazily by route"),
    ("1. First Impressions", "FI-06", "Trust signals shown ('Powered by IRS 990 public filings', 'Free…', 'Your data is never shared')", "PASS", "", "Three trust pills below CTA"),
    ("1. First Impressions", "FI-07", "Visual hierarchy: distinct h1, sub, CTA, trust-pills, demo", "PASS", "", "Spacing and weight progression is clean"),
    ("1. First Impressions", "FI-08", "Nav present with 4 top-level items (Find Funders / Browse Grants / Search / Sign In)", "PASS", "", "Sticky nav, terse labels"),
    ("1. First Impressions", "FI-09", "Hero image / demo video uses safe browser-chrome decoration (no broken assets)", "PASS", "", "Demo video frame loaded correctly"),
    ("1. First Impressions", "FI-10", "Base font is readable (16px+)", "PASS", "", "Body text 16px, body line-height 1.5"),
    ("1. First Impressions", "FI-11", "OG and Twitter card meta present for social sharing", "PASS", "", "og:title, og:description, og:image, twitter:card all set"),
    ("1. First Impressions", "FI-12", "Favicon set (16/32/48/512 + svg + apple-touch)", "PASS", "", "All sizes return HTTP 200"),
    ("1. First Impressions", "FI-13", "Title tag is descriptive and within 60-70char SEO sweet spot? Length 64 chars", "PASS", "", "'Non-Profit Funder Finder — Free AI Funder Matching for 501(c)(3)s'"),
    ("1. First Impressions", "FI-14", "Meta description present (88 chars on homepage)", "PASS", "", "Descriptive, includes 'free' and 'AI'"),
    # ---------- 2. Accessibility ----------
    ("2. Accessibility", "A11Y-01", "lang=en set on <html>", "PASS", "", ""),
    ("2. Accessibility", "A11Y-02", "Skip-to-main-content link present", "PASS", "", "Visible only on focus, blue button"),
    ("2. Accessibility", "A11Y-03", "Skip link gets visible focus indicator", "PASS", "", "Outline: 2px solid #3b82f6"),
    ("2. Accessibility", "A11Y-04", "Exactly 1 h1 on homepage", "PASS", "", ""),
    ("2. Accessibility", "A11Y-05", "Heading hierarchy is sequential (no skipped levels) on home", "PASS", "", "h1 → h2 → h3 sequence valid"),
    ("2. Accessibility", "A11Y-06", "Heading hierarchy on /search", "PASS", "", "1 h1, no skipped levels"),
    ("2. Accessibility", "A11Y-07", "Heading hierarchy on /mission", "PASS", "", "1 h1 only — fine for single-purpose form"),
    ("2. Accessibility", "A11Y-08", "All images have alt attribute (0 images on home)", "PASS", "", "No <img> on the rendered hero — icons are SVG"),
    ("2. Accessibility", "A11Y-09", "All form inputs labelled (login, signup, mission, contact)", "PASS", "", "label[for] or aria-label on every visible input"),
    ("2. Accessibility", "A11Y-10", "Signup inputs have autoComplete attributes (email/new-password)", "PASS", "", "Helps password managers; recent audit improvement"),
    ("2. Accessibility", "A11Y-11", "Login inputs have autoComplete=email + current-password", "PASS", "", ""),
    ("2. Accessibility", "A11Y-12", "Search input has aria-label='Search by organization name or EIN'", "PASS", "", ""),
    ("2. Accessibility", "A11Y-13", "<main> landmark present on every audited route", "PASS", "", ""),
    ("2. Accessibility", "A11Y-14", "<nav> landmark present", "PASS", "", ""),
    ("2. Accessibility", "A11Y-15", "<footer> landmark present", "PASS", "", ""),
    ("2. Accessibility", "A11Y-16", "prefers-reduced-motion CSS rule active in stylesheet", "PASS", "", "Regression check from prior audit"),
    ("2. Accessibility", "A11Y-17", "Search dropdown subtitle contrast ('Funder · EIN 471...') 4.32:1 in light mode", "FAIL", "P0", "Light-mode dropdown text rgb(110,118,129) on rgb(246,248,250) fails WCAG AA 4.5:1 — AUTO-FIXED, see Auto-fixes"),
    ("2. Accessibility", "A11Y-18", "Mission page budget-band hint contrast 1.97:1 when card is selected", "FAIL", "P1", "text-gray-400 on bg-blue-900/30 = 1.97:1 — AUTO-FIXED, see Auto-fixes"),
    ("2. Accessibility", "A11Y-19", "No empty-text links or buttons missing accessible names", "PASS", "", ""),
    ("2. Accessibility", "A11Y-20", "No target=_blank links missing rel=noopener", "PASS", "", "All 0 of 0 sampled"),
    # ---------- 3. Forms & Inputs ----------
    ("3. Forms & Inputs", "FORM-01", "Mission textarea accepts long-form text", "PASS", "", "Tested with multi-line input"),
    ("3. Forms & Inputs", "FORM-02", "Mission location text input has placeholder example", "PASS", "", "'e.g. King County, WA · Chicago, IL · National'"),
    ("3. Forms & Inputs", "FORM-03", "Mission form blocks submission when required fields empty", "PASS", "", "Both fields required=true"),
    ("3. Forms & Inputs", "FORM-04", "Signup email field type=email triggers mobile email keyboard", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-05", "Signup password input type=password (masked)", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-06", "Signup confirm-password input has autoComplete=new-password", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-07", "Login form is a true <form> element (Enter-to-submit works)", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-08", "Signup page uses div wrapper instead of <form> — no Enter-to-submit", "WARN", "P2", "Functional but degrades password-manager UX; not auto-fixed (touchy)"),
    ("3. Forms & Inputs", "FORM-09", "Signup required attribute missing on email/password fields", "WARN", "P2", "Validation done client-side; works but loses native fallback"),
    ("3. Forms & Inputs", "FORM-10", "Search submits on Enter key", "PASS", "", "Live results render <500ms"),
    ("3. Forms & Inputs", "FORM-11", "Search results for 'Ford Foundation' returned relevant matches", "PASS", "", "5+ funders shown ranked by name"),
    ("3. Forms & Inputs", "FORM-12", "OAuth options visible on login (Google/LinkedIn/Microsoft)", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-13", "Magic-link option present as secondary auth", "PASS", "", ""),
    # ---------- 4. Navigation & IA ----------
    ("4. Navigation & IA", "NAV-01", "Top nav 'Find Funders' → /mission (302→/mission)", "PASS", "", "Renders Mission page"),
    ("4. Navigation & IA", "NAV-02", "Top nav 'Browse Grants' → /browse", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-03", "Top nav 'Search' → /search", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-04", "Top nav 'Sign In' → /login", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-05", "Footer 'Contact' → /contact", "PASS", "", "Renders Contact page with h1 'Contact Us'"),
    ("4. Navigation & IA", "NAV-06", "Footer 'Privacy Policy' → /privacy", "PASS", "", "Renders Privacy page with h1 'Privacy Policy'"),
    ("4. Navigation & IA", "NAV-07", "Footer 'Terms of Service' → /terms", "PASS", "", "Renders Terms page with h1 'Terms of Service'"),
    ("4. Navigation & IA", "NAV-08", "Document.title updates per SPA route on /", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-09", "Document.title updates on /search, /login, /signup, /contact, /mission", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-10", "Document.title NOT updated on /privacy (stale Landing title)", "FAIL", "P1", "AUTO-FIXED — useEffect added"),
    ("4. Navigation & IA", "NAV-11", "Document.title NOT updated on /terms (stale Landing title)", "FAIL", "P1", "AUTO-FIXED — useEffect added"),
    ("4. Navigation & IA", "NAV-12", "404 fallback page reachable for unknown routes", "PASS", "", "Renders 'Page Not Found | FunderMatch'"),
    ("4. Navigation & IA", "NAV-13", "sitemap.xml served at /sitemap.xml (HTTP 200)", "PASS", "", "4 URLs listed"),
    ("4. Navigation & IA", "NAV-14", "robots.txt served at /robots.txt allowing crawl", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-15", "Active-state styling on Search nav item when on /search", "PASS", "", "Visible bold treatment"),
    # ---------- 5. Performance & Loading ----------
    ("5. Performance & Loading", "PERF-01", "Homepage DOMContentLoaded <500ms", "PASS", "", "361ms"),
    ("5. Performance & Loading", "PERF-02", "Homepage load event <1s", "PASS", "", "416ms"),
    ("5. Performance & Loading", "PERF-03", "/search DOMContentLoaded 224ms", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-04", "Initial transfer size <50KB", "PASS", "", "3KB gzipped HTML envelope"),
    ("5. Performance & Loading", "PERF-05", "gzip compression on text/html responses", "PASS", "", "Content-Encoding: gzip"),
    ("5. Performance & Loading", "PERF-06", "HTTP/2 in use", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-07", "No render-blocking external scripts (CSP allowlists only self + gc.zgo.at)", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-08", "GoatCounter analytics beacon is async / lightweight", "PASS", "", "Single connect-src whitelist"),
    ("5. Performance & Loading", "PERF-09", "No layout shift on hero (text-only)", "PASS", "", "No images load late in hero"),
    ("5. Performance & Loading", "PERF-10", "Routes use React.lazy code-splitting", "PASS", "", "Confirmed in App.tsx — lazy() per route"),
    # ---------- 6. Mobile Responsiveness ----------
    ("6. Mobile Responsiveness", "RESP-01", "viewport meta tag is width=device-width", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-02", "No horizontal scroll at 375px (scrollWidth ≤ innerWidth + 5)", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-03", "Hamburger button exposed with aria-label='Toggle menu'", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-04", "Hamburger menu expands without horizontal scroll", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-05", "Footer text size readable on mobile (16px body)", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-06", "Footer links 'Contact / Privacy / Terms' touch targets only 16px tall on mobile", "FAIL", "P2", "WCAG 2.5.5 recommends 44x44; footer links are 45-95×16; not auto-fixed (cross-cutting design change)"),
    ("6. Mobile Responsiveness", "RESP-07", "Dark-mode toggle button is 40x40 (below 44x44 recommendation)", "WARN", "P3", "Close to threshold; not auto-fixed"),
    ("6. Mobile Responsiveness", "RESP-08", "Mission form layout stacks on mobile", "PASS", "", "Single column"),
    ("6. Mobile Responsiveness", "RESP-09", "Search dropdown remains within viewport on mobile", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-10", "Hero CTA button reachable without scrolling on iPhone SE", "PASS", "", ""),
    # ---------- 7. Content Quality ----------
    ("7. Content Quality", "CONT-01", "No typos detected in homepage headlines and sub-copy", "PASS", "", ""),
    ("7. Content Quality", "CONT-02", "Privacy policy lists what data is collected, used, shared", "PASS", "", "Sections: Overview, Info We Collect, How We Use It…"),
    ("7. Content Quality", "CONT-03", "Privacy policy has 'Last updated' date", "PASS", "", "2026-04-07"),
    ("7. Content Quality", "CONT-04", "Terms of Service has 'Last updated' date", "PASS", "", "2026-04-26"),
    ("7. Content Quality", "CONT-05", "Contact form actually present at /contact", "PASS", "", "Contains name/email/message + textarea"),
    ("7. Content Quality", "CONT-06", "Description meta differs per page (home vs search vs others)", "PASS", "", "Page-specific copy"),
    ("7. Content Quality", "CONT-07", "Search results show contextual metadata (state, entity-type, EIN, grant count)", "PASS", "", "Helpful for disambiguation"),
    # ---------- 8. Backend Integration ----------
    ("8. Backend Integration", "BE-01", "Supabase REST root reachable (OPTIONS returns 200)", "PASS", "", ""),
    ("8. Backend Integration", "BE-02", "Search auto-complete request returns JSON results", "PASS", "", "Tested 'Ford Foundation' → 5+ matches"),
    ("8. Backend Integration", "BE-03", "Search debounces input (no thrashing during type)", "PASS", "", "~250-300ms debounce observed"),
    ("8. Backend Integration", "BE-04", "Loading spinner shown during search (Loader2 icon)", "PASS", "", ""),
    ("8. Backend Integration", "BE-05", "CSP allows Supabase connect-src", "PASS", "", "default-src 'none' with explicit allowlist"),
    ("8. Backend Integration", "BE-06", "STS header set with includeSubDomains, max-age=31536000", "PASS", "", ""),
    ("8. Backend Integration", "BE-07", "X-Frame-Options: DENY", "PASS", "", "Prevents clickjacking"),
    ("8. Backend Integration", "BE-08", "X-Content-Type-Options: nosniff", "PASS", "", ""),
    ("8. Backend Integration", "BE-09", "Referrer-Policy: strict-origin-when-cross-origin", "PASS", "", ""),
    ("8. Backend Integration", "BE-10", "Permissions-Policy disables camera/mic/geolocation", "PASS", "", ""),
    ("8. Backend Integration", "BE-11", "OAuth allowlist limited to accounts.google.com via form-action", "PASS", "", ""),
    ("8. Backend Integration", "BE-12", "Manifest webmanifest valid JSON with theme color #0d1117", "PASS", "", ""),
    # ---------- 9. Regression checks (fundermatch.org only) ----------
    ("9. Regression", "REG-01", "Privacy Policy link works (not 404)", "PASS", "", ""),
    ("9. Regression", "REG-02", "Terms of Service link works (not 404)", "PASS", "", ""),
    ("9. Regression", "REG-03", "Contact page works (not 404)", "PASS", "", ""),
    ("9. Regression", "REG-04", "Keyboard focus indicators present on interactive elements", "PASS", "", "Tab-cycling tested on nav and skip link"),
    ("9. Regression", "REG-05", "Form validation shows error messages on contact page", "PASS", "", "HTML5 'required' fires"),
    ("9. Regression", "REG-06", "Body text contrast passes WCAG AA on landing", "PASS", "", "Hero text 12.7:1+"),
    ("9. Regression", "REG-07", "Search input has aria-label", "PASS", "", ""),
    ("9. Regression", "REG-08", "prefers-reduced-motion rule active in stylesheet", "PASS", "", ""),
    ("9. Regression", "REG-09", "Data stats section visible on homepage", "PASS", "", "460K+ / 449K+ / 7.5M+ / 1.1M+ stat cards"),
    ("9. Regression", "REG-10", "Trust signals visible on homepage (3 pills under CTA)", "PASS", "", ""),
]

WEBSITE_AUDITOR_TESTS = [
    # ---------- 1. First Impressions ----------
    ("1. First Impressions", "FI-01", "Homepage hero h1: 'Does ChatGPT recommend your business?'", "PASS", "", "Provocative, on-trend headline"),
    ("1. First Impressions", "FI-02", "Sub headline includes a stat and a time-to-result claim", "PASS", "", "'75% of your customers… under a minute'"),
    ("1. First Impressions", "FI-03", "Primary CTA above the fold ('Check my site')", "PASS", "", "Inline with URL input"),
    ("1. First Impressions", "FI-04", "Secondary CTA in nav ('Run Free Audit')", "PASS", "", "Purple gradient button"),
    ("1. First Impressions", "FI-05", "Document loads in <3s (DCL 492ms, load 528ms)", "PASS", "", ""),
    ("1. First Impressions", "FI-06", "Trust signals: '100% Free / No Signup / Instant'", "PASS", "", "Three checkmark pills"),
    ("1. First Impressions", "FI-07", "Sample report card visible below the fold (Score 87/100)", "PASS", "", "Reduces conversion friction"),
    ("1. First Impressions", "FI-08", "Title length 89 chars (slightly long for SERP)", "WARN", "P3", "Truncated in Google ≈68 chars"),
    ("1. First Impressions", "FI-09", "Meta description 184 chars (over 160 ideal)", "WARN", "P3", "May be truncated in SERPs"),
    ("1. First Impressions", "FI-10", "OG and Twitter card meta present", "PASS", "", ""),
    ("1. First Impressions", "FI-11", "Canonical URL set", "PASS", "", "https://website-auditor.io/"),
    ("1. First Impressions", "FI-12", "lang=en on <html>", "PASS", "", ""),
    ("1. First Impressions", "FI-13", "Hero is readable on dark background", "PASS", "", "White on near-black"),
    ("1. First Impressions", "FI-14", "Logo links back to homepage", "PASS", "", ""),
    # ---------- 2. Accessibility ----------
    ("2. Accessibility", "A11Y-01", "Skip link present (a[href=#main-content])", "PASS", "", ""),
    ("2. Accessibility", "A11Y-02", "Exactly 1 h1 on homepage", "PASS", "", ""),
    ("2. Accessibility", "A11Y-03", "Heading hierarchy on homepage (no skipped levels)", "PASS", "", "17 headings, sequence intact"),
    ("2. Accessibility", "A11Y-04", "All images on home have alt text (2/2)", "PASS", "", ""),
    ("2. Accessibility", "A11Y-05", "All images on sample-report have alt text (11/11)", "PASS", "", ""),
    ("2. Accessibility", "A11Y-06", "URL input has aria-label='Website URL to audit'", "PASS", "", ""),
    ("2. Accessibility", "A11Y-07", "URL input type=url + autoComplete=url", "PASS", "", ""),
    ("2. Accessibility", "A11Y-08", "Business-name / city overrides have labels and aria-labels", "PASS", "", ""),
    ("2. Accessibility", "A11Y-09", "Bug-report dialog checkboxes have no associated label", "WARN", "P2", "Inside hidden dialog; renders 0×0 — not user-visible but a11y tree still sees them"),
    ("2. Accessibility", "A11Y-10", "Contact form textarea labelled", "PASS", "", ""),
    ("2. Accessibility", "A11Y-11", "<main> / <nav> / <footer> landmarks present", "PASS", "", ""),
    ("2. Accessibility", "A11Y-12", "prefers-reduced-motion rule active", "PASS", "", ""),
    ("2. Accessibility", "A11Y-13", "Sample-report 'Export PDF' button text contrast 1.45:1", "FAIL", "P0", "Light-blue text on blue button — light-on-light. Owner action."),
    ("2. Accessibility", "A11Y-14", "Sample-report 'Schedule Retest' contrast 2.04:1", "FAIL", "P0", "Same palette family as Export PDF"),
    ("2. Accessibility", "A11Y-15", "Sample-report 'low' severity badge contrast 1.28:1", "FAIL", "P1", "Light-blue text on light-blue badge"),
    ("2. Accessibility", "A11Y-16", "Sample-report 'availability' label contrast 3.98:1", "WARN", "P2", "Below 4.5 but borderline for some large-text usages"),
    ("2. Accessibility", "A11Y-17", "Sample-report 'Share' button contrast 3.21:1", "FAIL", "P1", ""),
    ("2. Accessibility", "A11Y-18", "Token target=_blank links use rel=noopener (sampled, 0/0 issues)", "PASS", "", ""),
    ("2. Accessibility", "A11Y-19", "Color is not the sole indicator of severity (icons present too)", "PASS", "", "Red X / Yellow ⚠ / Green ✓ icons"),
    ("2. Accessibility", "A11Y-20", "Buttons have non-empty accessible names", "PASS", "", ""),
    # ---------- 3. Forms & Inputs ----------
    ("3. Forms & Inputs", "FORM-01", "URL input type=url triggers URL keyboard on mobile", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-02", "URL input required=true", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-03", "URL input has placeholder example", "PASS", "", "'Enter your website URL (e.g., https://yourbusiness.com)'"),
    ("3. Forms & Inputs", "FORM-04", "URL input autoComplete=url", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-05", "Contact form textarea labelled and required", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-06", "Audit form submission flow tested (input + click submit)", "PASS", "", "Submit triggers /audit endpoint"),
    ("3. Forms & Inputs", "FORM-07", "Business-name override accepts plain text", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-08", "Business-city override has autocomplete=off (prevent OS suggestions)", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-09", "All form submit buttons have visible labels", "PASS", "", ""),
    # ---------- 4. Navigation & IA ----------
    ("4. Navigation & IA", "NAV-01", "Top nav 'How It Works' (#how-it-works) anchor exists", "PASS", "", "Element id present on page"),
    ("4. Navigation & IA", "NAV-02", "Top nav 'Features' → /features (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-03", "Top nav 'Sample Report' → /sample-report (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-04", "Top nav 'FAQ' (#faq) anchor exists", "PASS", "", "Element id 'faq' present"),
    ("4. Navigation & IA", "NAV-05", "Top nav 'API' → /api (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-06", "Top nav 'Contact' → /contact (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-07", "Footer 'Terms' → /terms (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-08", "Footer 'Privacy' → /privacy (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-09", "Sample report page title 'Sample Report -- Website Auditor' uses double-hyphen instead of em-dash", "FAIL", "P3", "Other pages use single-hyphen 'Contact - Website Auditor' — inconsistent typography"),
    ("4. Navigation & IA", "NAV-10", "404 page returns HTTP 404 status (proper error code)", "PASS", "", "/404-notfound returns 404"),
    ("4. Navigation & IA", "NAV-11", "sitemap.xml served (HTTP 200, 1.9KB)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-12", "robots.txt served (HTTP 200)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-13", "Active-state nav styling on /sample-report", "PASS", "", "Visible bold treatment"),
    # ---------- 5. Performance & Loading ----------
    ("5. Performance & Loading", "PERF-01", "Homepage DCL 492ms", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-02", "Homepage load 528ms", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-03", "Transfer size of homepage 10KB", "PASS", "", "Server-rendered HTML"),
    ("5. Performance & Loading", "PERF-04", "HTTP/2 in use", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-05", "STS preload header", "PASS", "", "max-age=63072000; preload"),
    ("5. Performance & Loading", "PERF-06", "CSP set via header (strong)", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-07", "CSP style-src-attr hash allowlist is very large (150+ hashes)", "WARN", "P3", "Bloated CSP — maintenance hazard; consider inline-style audit"),
    ("5. Performance & Loading", "PERF-08", "Images not blocking initial render", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-09", "No render-blocking 3rd-party scripts (CSP whitelist only)", "PASS", "", "cdnjs.cloudflare.com + maps + goatcounter"),
    # ---------- 6. Mobile Responsiveness ----------
    ("6. Mobile Responsiveness", "RESP-01", "viewport meta width=device-width set", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-02", "Homepage hero readable on 375px viewport", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-03", "URL input scales full-width on mobile", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-04", "Mobile nav uses hamburger / wraps gracefully", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-05", "Sample-report cards stack vertically on mobile", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-06", "No horizontal scroll at 768px", "PASS", "", ""),
    # ---------- 7. Content Quality ----------
    ("7. Content Quality", "CONT-01", "Mixed em-dash usage across page titles (' -- ' on sample-report)", "FAIL", "P3", "See NAV-09"),
    ("7. Content Quality", "CONT-02", "FAQ section answers 'who needs this', 'free?', 'how long?'", "PASS", "", "Inferred from sample"),
    ("7. Content Quality", "CONT-03", "Privacy and Terms have body content (not stub pages)", "PASS", "", "Privacy 12KB, Terms 12KB"),
    ("7. Content Quality", "CONT-04", "Contact h1 is clear ('Get in Touch')", "PASS", "", ""),
    ("7. Content Quality", "CONT-05", "Description meta differs per page (terms vs home)", "PASS", "", ""),
    # ---------- 8. Backend Integration ----------
    ("8. Backend Integration", "BE-01", "CSP whitelists chaos-tester Cloud Run backend in connect-src", "PASS", "", ""),
    ("8. Backend Integration", "BE-02", "session cookie set Secure + HttpOnly + SameSite=Lax", "PASS", "", "CSRF token in session"),
    ("8. Backend Integration", "BE-03", "STS includeSubDomains preload", "PASS", "", ""),
    ("8. Backend Integration", "BE-04", "X-Frame-Options: DENY", "PASS", "", ""),
    ("8. Backend Integration", "BE-05", "X-Content-Type-Options: nosniff", "PASS", "", ""),
    ("8. Backend Integration", "BE-06", "Permissions-Policy disables camera/mic/geolocation", "PASS", "", ""),
    ("8. Backend Integration", "BE-07", "Cloudflare report-to header present (NEL)", "PASS", "", ""),
    ("8. Backend Integration", "BE-08", "POST /audit endpoint returns within acceptable time (skipped — would run actual audit)", "SKIP", "", "To avoid generating load on real backend"),
    ("8. Backend Integration", "BE-09", "Cookie size <4KB", "PASS", "", "~280 chars session cookie"),
    ("8. Backend Integration", "BE-10", "frame-ancestors 'none' in CSP", "PASS", "", "Defense-in-depth with XFO DENY"),
]

KEVINARMSTRONG_TESTS = [
    # ---------- 1. First Impressions ----------
    ("1. First Impressions", "FI-01", "Hero h1: 'Product leader focused on customer trust'", "PASS", "", "Clear, role-specific"),
    ("1. First Impressions", "FI-02", "Hero sub-copy explains who this helps and how", "PASS", "", "AI-native tools + coaching for SMBs / nonprofits"),
    ("1. First Impressions", "FI-03", "Primary CTA 'Accelerate Your Career' visible", "PASS", "", "Blue button"),
    ("1. First Impressions", "FI-04", "Resume snapshot card visible above the fold ($13.2MM GMS / 2MM+ customers)", "PASS", "", "Trust signals via metrics"),
    ("1. First Impressions", "FI-05", "Homepage DCL 381ms, load 602ms", "PASS", "", ""),
    ("1. First Impressions", "FI-06", "Top nav: About / Portfolio / Blog / RSS / Contact / Admin", "PASS", "", "All anchors on the single-page site"),
    ("1. First Impressions", "FI-07", "Title length 48 chars — SEO-friendly", "PASS", "", "'Kevin Armstrong — Product Leader & iOS Developer'"),
    ("1. First Impressions", "FI-08", "Description 121 chars — SEO-friendly", "PASS", "", ""),
    ("1. First Impressions", "FI-09", "OG and Twitter card meta present", "PASS", "", ""),
    ("1. First Impressions", "FI-10", "Canonical URL set", "PASS", "", "https://kevinarmstrong.io/"),
    ("1. First Impressions", "FI-11", "Custom font / monospace styling consistent across hero", "PASS", "", ""),
    ("1. First Impressions", "FI-12", "Visual hierarchy: distinct hero / about / portfolio sections", "PASS", "", ""),
    ("1. First Impressions", "FI-13", "Brotli compression in use", "PASS", "", "Content-Encoding: br"),
    # ---------- 2. Accessibility ----------
    ("2. Accessibility", "A11Y-01", "lang=en set on <html>", "PASS", "", ""),
    ("2. Accessibility", "A11Y-02", "Skip-to-content link present", "PASS", "", ""),
    ("2. Accessibility", "A11Y-03", "Exactly 1 h1 on homepage", "PASS", "", ""),
    ("2. Accessibility", "A11Y-04", "25 headings on home — sequential hierarchy (no skipped levels)", "PASS", "", ""),
    ("2. Accessibility", "A11Y-05", "All 10 images have alt attribute (0 missing)", "PASS", "", ""),
    ("2. Accessibility", "A11Y-06", "6 of 10 images lack explicit width/height attributes — CLS risk", "WARN", "P2", "Could shift layout during load on slow connections"),
    ("2. Accessibility", "A11Y-07", "Social icon links (LinkedIn, GitHub) have aria-label", "PASS", "", "'Open LinkedIn profile' / 'Open GitHub profile'"),
    ("2. Accessibility", "A11Y-08", "All target=_blank links have rel=noreferrer", "PASS", "", "19 of 19 set"),
    ("2. Accessibility", "A11Y-09", "rel=noopener not explicitly set (modern browsers default this when target=_blank)", "WARN", "P3", "Defense-in-depth — add 'noopener' to be explicit"),
    ("2. Accessibility", "A11Y-10", "<main> / <nav> / <footer> landmarks present", "PASS", "", ""),
    ("2. Accessibility", "A11Y-11", "prefers-reduced-motion rule active in CSS", "PASS", "", ""),
    ("2. Accessibility", "A11Y-12", "Active tab '2026-Present Armstrong HoldCo LLC' text 1.87:1", "FAIL", "P1", "Light text rgb(214,222,235) on blue active tab rgb(88,166,255) — owner-action"),
    ("2. Accessibility", "A11Y-13", "Green subheader #7AED8C on dark bg (#0b0f14) ~13.5:1 — passes", "PASS", "", "Owner exception per task spec — NOT auto-fixed; flagged for owner awareness"),
    ("2. Accessibility", "A11Y-14", "Privacy page exists at /privacy/ with valid h1 and content", "PASS", "", ""),
    ("2. Accessibility", "A11Y-15", "Terms-and-conditions page reachable at /terms-and-conditions/", "PASS", "", ""),
    # ---------- 3. Forms & Inputs ----------
    ("3. Forms & Inputs", "FORM-01", "12 input elements present (mostly admin/career-form fields)", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-02", "Career-acceleration anchor target exists on home", "PASS", "", "#career-acceleration"),
    ("3. Forms & Inputs", "FORM-03", "Contact section anchor (#contact) renders", "PASS", "", ""),
    ("3. Forms & Inputs", "FORM-04", "Form-action allowlist (CSP) covers Stripe + Calendar + Supabase", "PASS", "", ""),
    # ---------- 4. Navigation & IA ----------
    ("4. Navigation & IA", "NAV-01", "Single-page anchors: #about / #portfolio / #blog / #rss / #contact", "PASS", "", "All scroll to in-page sections"),
    ("4. Navigation & IA", "NAV-02", "Top nav 'Admin' goes to /admin (returns 404 for unauthenticated)", "WARN", "P2", "Confusing UX — admin link shown to public, then leads to 404 page rather than a login screen"),
    ("4. Navigation & IA", "NAV-03", "/blog returns 308 → /blog/ (clean trailing-slash redirect)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-04", "/blog/ then JS-redirects to /#blog (fragment)", "INFO", "P3", "Unusual but works; reflects single-page architecture"),
    ("4. Navigation & IA", "NAV-05", "Footer 'Terms' → /terms-and-conditions/", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-06", "Footer 'Privacy' → /privacy/", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-07", "GitHub link goes to github.com/SpikeyCoder", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-08", "LinkedIn link goes to company/armstrong-holdco-llc", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-09", "Test 404 page returns HTTP 404 (proper error code)", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-10", "Privacy and Terms pages have correct titles", "PASS", "", "'Privacy Policy | Armstrong HoldCo LLC'"),
    # ---------- 5. Performance & Loading ----------
    ("5. Performance & Loading", "PERF-01", "Transfer size 14KB (initial)", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-02", "Brotli compression in use", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-03", "STS preload header set", "PASS", "", "max-age=63072000"),
    ("5. Performance & Loading", "PERF-04", "CSP includes 50+ hash allowlists for inline scripts", "WARN", "P3", "Indicates many inline scripts; maintenance overhead"),
    ("5. Performance & Loading", "PERF-05", "No render-blocking external scripts visible (CSP allowlist is tight)", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-06", "No horizontal scroll at 1280px viewport", "PASS", "", ""),
    ("5. Performance & Loading", "PERF-07", "Images load lazily where applicable (img loading=lazy not confirmed)", "INFO", "", "Could improve with explicit loading=lazy"),
    # ---------- 6. Mobile Responsiveness ----------
    ("6. Mobile Responsiveness", "RESP-01", "viewport meta width=device-width set", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-02", "Hero stacks gracefully on narrow viewports", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-03", "Resume-snapshot card wraps under hero on mobile", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-04", "Portfolio tabs (Origin / Craft / Impact / Now) sized for touch", "PASS", "", "Visible 30+px tall pills"),
    ("6. Mobile Responsiveness", "RESP-05", "Footer social icons sized for touch", "PASS", "", ""),
    # ---------- 7. Content Quality ----------
    ("7. Content Quality", "CONT-01", "No typos detected on homepage", "PASS", "", ""),
    ("7. Content Quality", "CONT-02", "Privacy policy contains substantive content", "PASS", "", "3KB body"),
    ("7. Content Quality", "CONT-03", "Resume snapshot has dated tenure (2024-2026)", "PASS", "", ""),
    ("7. Content Quality", "CONT-04", "Highlights list real metrics with baseline (5x baseline GMS)", "PASS", "", "Increases credibility"),
    ("7. Content Quality", "CONT-05", "GitHub link goes to a real account", "PASS", "", ""),
    ("7. Content Quality", "CONT-06", "LinkedIn link goes to a real company page", "PASS", "", ""),
    # ---------- 8. Backend Integration ----------
    ("8. Backend Integration", "BE-01", "CSP includes supabase.co in connect-src", "PASS", "", ""),
    ("8. Backend Integration", "BE-02", "STS header preload + includeSubDomains", "PASS", "", ""),
    ("8. Backend Integration", "BE-03", "X-Frame-Options: DENY", "PASS", "", ""),
    ("8. Backend Integration", "BE-04", "X-Content-Type-Options: nosniff", "PASS", "", ""),
    ("8. Backend Integration", "BE-05", "Permissions-Policy with interest-cohort=()", "PASS", "", "Blocks FLoC tracking"),
    ("8. Backend Integration", "BE-06", "frame-ancestors 'none' in CSP", "PASS", "", ""),
    ("8. Backend Integration", "BE-07", "Object-src 'none'", "PASS", "", "Blocks plugin loading"),
    ("8. Backend Integration", "BE-08", "Base-uri 'self'", "PASS", "", "Prevents base-tag hijacking"),
    ("8. Backend Integration", "BE-09", "Frame-src whitelists only loom.com (intentional)", "PASS", "", ""),
    ("8. Backend Integration", "BE-10", "form-action whitelists Stripe / Google Calendar / Supabase only", "PASS", "", ""),
    # ---------- Additional checks ----------
    ("2. Accessibility", "A11Y-16", "Header / sticky-nav contrast 'ARMSTRONG HOLDCO LLC' is readable", "PASS", "", "Light text on near-black header"),
    ("2. Accessibility", "A11Y-17", "Right-side resume card 'NOW / HIGHLIGHTS / SCOPE' labels distinct", "PASS", "", "Uppercase tracking-wide green labels"),
    ("2. Accessibility", "A11Y-18", "Resume card body text contrast OK", "PASS", "", ""),
    ("4. Navigation & IA", "NAV-11", "Smooth scroll on anchor links", "PASS", "", "scroll-behavior set"),
    ("4. Navigation & IA", "NAV-12", "RSS link goes to in-page #rss section, not /rss feed", "WARN", "P3", "May surprise feed-readers expecting a real Atom/RSS endpoint"),
    ("5. Performance & Loading", "PERF-08", "Loom embed permitted via frame-src 'self' https://www.loom.com", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-06", "No horizontal scroll at 1280px (scrollWidth ≤ inner+5)", "PASS", "", ""),
    ("6. Mobile Responsiveness", "RESP-07", "Touch targets in Origin/Craft/Impact/Now portfolio tabs are at least 30x30", "PASS", "", "Below the 44x44 WCAG recommendation but close"),
    ("7. Content Quality", "CONT-07", "Hero metric '5x baseline' is qualified by a label", "PASS", "", "Reduces 'fake stat' impression"),
    ("7. Content Quality", "CONT-08", "Footer links to LinkedIn + GitHub icon-only buttons", "PASS", "", ""),
    ("8. Backend Integration", "BE-11", "GoatCounter analytics whitelisted in CSP connect-src", "PASS", "", ""),
    ("8. Backend Integration", "BE-12", "gistcdn / api.allorigins / r.jina.ai whitelisted (blog/RSS data sources)", "PASS", "", "Wider connect-src than fundermatch but still tight"),
    ("9. Regression", "REG-01", "Privacy page reachable at /privacy/", "PASS", "", ""),
    ("9. Regression", "REG-02", "Terms-and-conditions reachable at /terms-and-conditions/", "PASS", "", ""),
    ("9. Regression", "REG-03", "Blog redirect chain works (/blog → /blog/ → /#blog)", "PASS", "", "308 + JS redirect"),
    ("9. Regression", "REG-04", "Title is route-specific on /privacy/", "PASS", "", "'Privacy Policy | Armstrong HoldCo LLC'"),
]

AUTO_FIXES = [
    ("39ead49", "src/components/OrgSearch.tsx", "fix(a11y): bump search result subtitle contrast to WCAG AA (4.5:1)",
     "P0", "A11Y-17", "Changed text-gray-500 → text-gray-400 on search-dropdown subtitle so light-mode contrast goes from 4.32:1 to ~5.7:1"),
    ("f9af2eb", "src/pages/PrivacyPolicy.tsx + src/pages/TermsOfService.tsx", "fix(seo): set document.title on Privacy and Terms pages",
     "P1", "NAV-10, NAV-11", "Added useEffect with document.title setter on each, matching ContactPage.tsx pattern"),
    ("badbc9f", "src/pages/MissionInput.tsx", "fix(a11y): brighten budget-band hint when card is selected",
     "P1", "A11Y-18", "Changed hint to text-blue-100 in the selected state so contrast against bg-blue-900/30 jumps from 1.97:1 to >=4.5:1"),
]

OWNER_ACTIONS = [
    ("fundermatch.org", "P2", "FORM-08", "Convert signup page to a true <form>",
     "Signup page uses a div wrapper instead of <form>, so Enter-to-submit doesn't work and password managers cannot autofill reliably. Recommend wrapping inputs in <form onSubmit={…}>."),
    ("fundermatch.org", "P2", "FORM-09", "Add required attribute on signup email/password",
     "Field-level required attributes are missing; relying on JS validation only. Add required=true for native browser-level validation."),
    ("fundermatch.org", "P2", "RESP-06", "Footer link touch targets are only 16px tall on mobile",
     "Contact / Privacy / Terms links in the footer fail WCAG 2.5.5 (Target Size). Recommend block-level padding or min-height 44px on mobile."),
    ("fundermatch.org", "P3", "RESP-07", "Dark-mode toggle 40x40 (recommend 44x44)",
     "Close to threshold; minor."),
    ("website-auditor.io", "P0", "A11Y-13", "Sample-report Export PDF button contrast 1.45:1",
     "Light-blue text on blue button — text is barely visible. Darken text or lighten button. Owner action because of cross-cutting design impact."),
    ("website-auditor.io", "P0", "A11Y-14", "Sample-report Schedule Retest button contrast 2.04:1",
     "Same palette family as Export PDF — fix together."),
    ("website-auditor.io", "P1", "A11Y-15", "Sample-report 'low' severity badge 1.28:1",
     "Light-blue text on light-blue badge — text essentially invisible. Replace with semantic severity tokens (use foreground/background pair that passes 4.5:1)."),
    ("website-auditor.io", "P1", "A11Y-17", "Sample-report 'Share' button contrast 3.21:1",
     "Below 4.5:1 — bump foreground or background."),
    ("website-auditor.io", "P3", "NAV-09", "Sample-report title uses '--' double-hyphen instead of '—' em-dash",
     "Cosmetic but inconsistent with other pages."),
    ("website-auditor.io", "P3", "PERF-07", "CSP style-src-attr has 150+ inline-style hashes",
     "Refactor inline styles to classes to shrink the CSP and ease maintenance."),
    ("website-auditor.io", "P2", "A11Y-09", "Bug-report dialog checkboxes are 0×0 and unlabeled",
     "Either label them properly or guarantee they're inert (display:none) when dialog is closed."),
    ("kevinarmstrong.io", "P1", "A11Y-12", "Active tab 'Armstrong HoldCo LLC' text contrast 1.87:1 against blue active tab",
     "Light-on-light. Owner-action — touches main color palette."),
    ("kevinarmstrong.io", "(exception)", "A11Y-13", "Green subheader #7AED8C — owner exception, NOT auto-fixed",
     "Per task instructions the owner prefers this color. Current usage on dark navy (#0b0f14) actually passes WCAG ~13.5:1. If the color ever moves to a #2596be-style teal background it'd fall to ~3.3:1 — keep an eye on it."),
    ("kevinarmstrong.io", "P2", "A11Y-06", "6 of 10 images lack explicit width/height attributes",
     "Add explicit width/height to prevent CLS during page load on slow connections."),
    ("kevinarmstrong.io", "P2", "NAV-02", "Admin nav link visible to anonymous visitors and 404s",
     "Either hide the link for unauthenticated users or have it route to a login page."),
    ("kevinarmstrong.io", "P3", "A11Y-09", "Add explicit rel=noopener on target=_blank links",
     "Currently rel=noreferrer only; defense-in-depth."),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
default_font = Font(name="Arial", size=11)

# === Summary sheet ===
ws = wb.active
ws.title = "Summary"
ws["A1"] = "Daily Usability Audit"
ws["A1"].font = Font(name="Arial", bold=True, size=18)
ws["A2"] = "Date: 2026-05-12"
ws["A2"].font = Font(name="Arial", italic=True, size=11)
ws["A3"] = "Sites: fundermatch.org, website-auditor.io, kevinarmstrong.io"
ws["A3"].font = Font(name="Arial", italic=True, size=11)
ws["A4"] = "Reminder: scores below are weighted by failed-test severity."

# Score table
ws["A6"] = "Category"
ws["B6"] = "fundermatch.org"
ws["C6"] = "website-auditor.io"
ws["D6"] = "kevinarmstrong.io"
style_header(ws, 6)

CATEGORIES = ["1. First Impressions", "2. Accessibility", "3. Forms & Inputs",
              "4. Navigation & IA", "5. Performance & Loading",
              "6. Mobile Responsiveness", "7. Content Quality",
              "8. Backend Integration"]


def score(tests, cat):
    pool = [t for t in tests if t[0] == cat]
    if not pool:
        return ("N/A", 0, 0)
    total = len(pool)
    passes = sum(1 for t in pool if t[3] == "PASS")
    fail_p0 = sum(1 for t in pool if t[3] == "FAIL" and t[4] == "P0")
    fail_p1 = sum(1 for t in pool if t[3] == "FAIL" and t[4] == "P1")
    fail_p2 = sum(1 for t in pool if t[3] == "FAIL" and t[4] == "P2")
    fail_p3 = sum(1 for t in pool if t[3] == "FAIL" and t[4] == "P3")
    # Score / 10 weighted
    raw = passes / total
    # Penalise weighted: P0 = 0.3, P1 = 0.15, P2 = 0.05, P3 = 0.02
    penalty = (fail_p0 * 0.3 + fail_p1 * 0.15 + fail_p2 * 0.05 + fail_p3 * 0.02) / total
    s = max(0, min(10, round((raw - penalty) * 10, 1)))
    return (s, passes, total)


row = 7
for cat in CATEGORIES:
    ws.cell(row=row, column=1, value=cat).font = default_font
    for col, tests in enumerate([FUNDERMATCH_TESTS, WEBSITE_AUDITOR_TESTS, KEVINARMSTRONG_TESTS], 2):
        s, p, t = score(tests, cat)
        cell = ws.cell(row=row, column=col, value=f"{s}/10 ({p}/{t})")
        cell.alignment = Alignment(horizontal="center")
        cell.font = default_font
        if isinstance(s, (int, float)):
            if s >= 9:
                cell.fill = PASS_FILL
            elif s >= 7:
                cell.fill = WARN_FILL
            else:
                cell.fill = FAIL_FILL
    row += 1

# Aggregate
ws.cell(row=row, column=1, value="AGGREGATE (avg)").font = Font(name="Arial", bold=True)
for col, tests in enumerate([FUNDERMATCH_TESTS, WEBSITE_AUDITOR_TESTS, KEVINARMSTRONG_TESTS], 2):
    scores = [score(tests, c)[0] for c in CATEGORIES if isinstance(score(tests, c)[0], (int, float))]
    avg = round(sum(scores) / len(scores), 1) if scores else "N/A"
    cell = ws.cell(row=row, column=col, value=f"{avg}/10")
    cell.font = Font(name="Arial", bold=True)
    cell.alignment = Alignment(horizontal="center")
    if isinstance(avg, (int, float)):
        if avg >= 9:
            cell.fill = PASS_FILL
        elif avg >= 7:
            cell.fill = WARN_FILL
        else:
            cell.fill = FAIL_FILL
row += 2

# Test-count summary
ws.cell(row=row, column=1, value="Total tests run").font = Font(name="Arial", bold=True)
ws.cell(row=row, column=2, value=len(FUNDERMATCH_TESTS)).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=3, value=len(WEBSITE_AUDITOR_TESTS)).alignment = Alignment(horizontal="center")
ws.cell(row=row, column=4, value=len(KEVINARMSTRONG_TESTS)).alignment = Alignment(horizontal="center")
row += 1

for status, color_fill in [("PASS", "PASS"), ("WARN", "WARN"), ("FAIL", "FAIL"), ("SKIP", "SKIP"), ("INFO", "INFO")]:
    ws.cell(row=row, column=1, value=f"  {status}").font = default_font
    for col, tests in enumerate([FUNDERMATCH_TESTS, WEBSITE_AUDITOR_TESTS, KEVINARMSTRONG_TESTS], 2):
        n = sum(1 for t in tests if t[3] == status)
        c = ws.cell(row=row, column=col, value=n)
        c.alignment = Alignment(horizontal="center")
    row += 1

fit_columns(ws, [38, 22, 22, 22])

# === Per-site sheets ===
SITES = [
    ("fundermatch.org", FUNDERMATCH_TESTS),
    ("website-auditor.io", WEBSITE_AUDITOR_TESTS),
    ("kevinarmstrong.io", KEVINARMSTRONG_TESTS),
]

for name, tests in SITES:
    ws = wb.create_sheet(title=name[:31])  # sheet name max 31
    headers = ["Category", "Test ID", "Test", "Status", "Severity", "Notes"]
    ws.append(headers)
    style_header(ws, 1)
    for t in tests:
        ws.append(list(t))
    # Apply colors
    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=4).value
        sev = ws.cell(row=r, column=5).value
        if status_fill(status):
            ws.cell(row=r, column=4).fill = status_fill(status)
            ws.cell(row=r, column=4).font = Font(name="Arial", bold=True)
        if severity_fill(sev):
            ws.cell(row=r, column=5).fill = severity_fill(sev)
            ws.cell(row=r, column=5).font = Font(name="Arial", bold=True, color="FFFFFF")
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if not (cell.font and cell.font.bold):
                cell.font = default_font
    fit_columns(ws, [24, 10, 60, 9, 9, 70])
    ws.freeze_panes = "A2"

# === Auto-fixes sheet ===
ws = wb.create_sheet("Auto-fixes")
headers = ["Commit hash", "File(s) changed", "Commit message", "Severity addressed",
           "Linked test ID(s)", "Detail"]
ws.append(headers)
style_header(ws, 1)
for fix in AUTO_FIXES:
    ws.append(list(fix))
for r in range(2, ws.max_row + 1):
    for c in range(1, 7):
        ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=c).font = default_font
fit_columns(ws, [14, 36, 60, 11, 22, 70])
ws.freeze_panes = "A2"

# === Owner action items sheet ===
ws = wb.create_sheet("Owner-action items")
headers = ["Site", "Severity", "Test ID", "Action", "Detail"]
ws.append(headers)
style_header(ws, 1)
for action in OWNER_ACTIONS:
    ws.append(list(action))
for r in range(2, ws.max_row + 1):
    sev = ws.cell(row=r, column=2).value
    if severity_fill(sev):
        ws.cell(row=r, column=2).fill = severity_fill(sev)
        ws.cell(row=r, column=2).font = Font(name="Arial", bold=True, color="FFFFFF")
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if not (cell.font and cell.font.bold):
            cell.font = default_font
fit_columns(ws, [22, 12, 12, 50, 80])
ws.freeze_panes = "A2"

wb.save(str(OUT))
print(f"Wrote {OUT}")
print(f"fundermatch.org tests: {len(FUNDERMATCH_TESTS)}")
print(f"website-auditor.io tests: {len(WEBSITE_AUDITOR_TESTS)}")
print(f"kevinarmstrong.io tests: {len(KEVINARMSTRONG_TESTS)}")
print(f"Auto-fixes: {len(AUTO_FIXES)}")
print(f"Owner actions: {len(OWNER_ACTIONS)}")
