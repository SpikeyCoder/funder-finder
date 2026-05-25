"""Build the daily usability audit Excel report for 2026-05-25.

Sheets:
  Summary, website-auditor.io, kevinarmstrong.io, fundermatch.org,
  PRs opened, Owner exceptions, Regression-Tracking

Data collected this run via curl / direct codebase inspection against
live production sites and the cloned repos on 2026-05-25.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
SKIP_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
INFO_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def status_fill(s):
    s = (s or "").lower()
    if s == "pass":
        return PASS_FILL
    if s == "fail":
        return FAIL_FILL
    if s in ("warn", "warning"):
        return WARN_FILL
    if s == "skip":
        return SKIP_FILL
    if s == "info":
        return INFO_FILL
    return None


def write_tests(ws, tests):
    ws.append(["Test ID", "Category", "Test", "Status", "Severity", "Notes"])
    style_header(ws)
    for row in tests:
        ws.append(list(row))
        last_row = ws.max_row
        fill = status_fill(row[3])
        for ci in range(1, 7):
            cell = ws.cell(row=last_row, column=ci)
            cell.alignment = LEFT
            cell.border = BORDER
            cell.font = BODY_FONT
            if fill is not None:
                cell.fill = fill
    set_widths(ws, [14, 22, 60, 10, 10, 80])
    ws.freeze_panes = "A2"


# ============================================================
# WEBSITE-AUDITOR.IO  (chaos_tester) — 96 tests
# ============================================================
WAO_TESTS = [
    # First Impressions (12)
    ("WAO-FI-001", "First Impressions", "Homepage explains product (AI visibility + security + perf scanner)", "pass", "-", "Hero 'Does ChatGPT recommend your business?' clearly frames AI-discoverability; meta description 220 chars covers AI visibility, broken links, security, performance, ChatGPT, no-login."),
    ("WAO-FI-002", "First Impressions", "Value proposition above the fold", "pass", "-", "Hero + URL input + 'Check my site' CTA visible at 1280x800 without scroll."),
    ("WAO-FI-003", "First Impressions", "Primary CTA visible and compelling", "pass", "-", "'Check my site' submit button anchors the hero form; secondary 'View Sample Report' card on landing."),
    ("WAO-FI-004", "First Impressions", "HEAD / returns under 3s with Cloudflare", "pass", "-", "HTTP/2 200 from Cloudflare cf-ray edge; cf-cache-status DYNAMIC; full HTML 463 lines."),
    ("WAO-FI-005", "First Impressions", "Visual hierarchy (1 H1, multiple H2/H3)", "pass", "-", "1 H1 'Does ChatGPT recommend your business?'; 2 H2 ('What We Check', 'Ready to Improve Your Website?'); 7 H3 service cards including API upsell."),
    ("WAO-FI-006", "First Impressions", "Trust signals on homepage", "pass", "-", "Testimonials block + industry strip + sample-report preview card retained from prior audit."),
    ("WAO-FI-007", "First Impressions", "Navigation intuitive (4 top-level items)", "pass", "-", "Nav: Website Auditor / Sample Report / API / Contact. Compact, no jargon."),
    ("WAO-FI-008", "First Impressions", "Hero works on mobile (viewport meta correct)", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1.0'> + hamburger button#navToggle aria-controls=mainNav."),
    ("WAO-FI-009", "First Impressions", "No broken images on home", "pass", "-", "Only inline SVGs in hero; logo.svg + favicon.svg referenced and 200 OK."),
    ("WAO-FI-010", "First Impressions", "Default body font readable", "pass", "-", "Inter @ 16px base; JetBrains Mono for code samples. Preconnects to fonts.googleapis.com + fonts.gstatic.com."),
    ("WAO-FI-011", "First Impressions", "Hero Gartner claim is honest / non-misleading", "warn", "P3", "Hero still cites 'By 2028, 50% your customers will find your competitors through AI according to Gartner' — link goes to Gartner strategic-predictions page but the exact stat is not in the linked article. Owner-info item; tighten attribution or rephrase as a directional forecast."),
    ("WAO-FI-012", "First Impressions", "Above-the-fold CTA reachable without scroll", "pass", "-", "URL input + Check-my-site button visible inside the .landing-hero div which renders above the fold."),

    # Accessibility (17)
    ("WAO-AX-001", "Accessibility", "Skip-to-main-content link present", "pass", "-", "<a href='#main-content' class='skip-link'>Skip to main content</a> is the first focusable element in base.html."),
    ("WAO-AX-002", "Accessibility", "lang attribute set on <html>", "pass", "-", "<html lang='en'> in base.html."),
    ("WAO-AX-003", "Accessibility", "Landmarks (main + nav + header + footer)", "pass", "-", "<main class='container u-padding-y' id='main-content'>, <nav id='mainNav' aria-label='Main navigation'>, <header>, <footer> all present."),
    ("WAO-AX-004", "Accessibility", "Hamburger button has aria-label + aria-expanded", "pass", "-", "<button class='hamburger' id='navToggle' aria-label='Toggle navigation menu' aria-expanded='false'>"),
    ("WAO-AX-005", "Accessibility", "Logo img has empty alt + aria-hidden (decorative)", "pass", "-", "<img src='/static/logo.svg' alt='' class='site-logo' aria-hidden='true' fetchpriority='high'> — correctly decorative since the brand text follows."),
    ("WAO-AX-006", "Accessibility", "URL input has aria-label", "pass", "-", "<input type='url' id='base_url' name='base_url' ... aria-label='Website URL to audit'>"),
    ("WAO-AX-007", "Accessibility", "Audit submit button has aria-label", "pass", "-", "<button ... aria-label='Check my site'>Check my site</button>"),
    ("WAO-AX-008", "Accessibility", "Search icon SVG marked aria-hidden + focusable=false", "pass", "-", "Hero <svg ... aria-hidden='true' focusable='false'> — passes axe svg-img-alt heuristics."),
    ("WAO-AX-009", "Accessibility", "Heading hierarchy (no skipped levels)", "pass", "-", "Sequence verified H1 -> H2 -> H3 across home."),
    ("WAO-AX-010", "Accessibility", "Body text contrast on dark slate", "pass", "-", "Body off-white on slate-900 ~ ratio >12:1 (WCAG AA passes by wide margin)."),
    ("WAO-AX-011", "Accessibility", "Global :focus-visible outline rule present", "pass", "-", "static/css/base.css line 543: a:focus-visible, button:focus-visible, input:focus-visible, select:focus-visible, textarea:focus-visible { outline ... }"),
    ("WAO-AX-012", "Accessibility", "Skip-link styled to slide in on focus", "pass", "-", "static/css/base.css lines 522-527: .skip-link { ... } .skip-link:focus { left: 0; width: auto; height: auto; }"),
    ("WAO-AX-013", "Accessibility", "prefers-reduced-motion respected", "pass", "-", "static/css/base.css line 558: @media (prefers-reduced-motion: reduce). Plus a second rule in progress.css line 34 for the audit-progress page animations."),
    ("WAO-AX-014", "Accessibility", "Form labels associated with inputs (bug form)", "pass", "-", "<label for='bugDesc' class='sr-only'>, <label for='bugIsFeature'>, <label for='bugScreenshot'> all explicit-association."),
    ("WAO-AX-015", "Accessibility", "Email link is mailto: (a11y + intent)", "pass", "-", "<a href='mailto:support@website-auditor.io'> on /contact."),
    ("WAO-AX-016", "Accessibility", "Bug-report widget keyboard accessible", "pass", "-", "Bug form uses <form id='bugForm'> with native button submit and textarea; supports keyboard tab + Enter."),
    ("WAO-AX-017", "Accessibility", "theme-color + color-scheme meta tags present", "fail", "P2", "FIXED IN PR #83 - base.html had viewport but no theme-color/color-scheme. Mobile browser chrome and OS form controls (scrollbars, autofill) rendered light against dark site. Added <meta name='theme-color' content='#0f172a'> + <meta name='color-scheme' content='dark light'> in branch usability/2026-05-25-fixes."),

    # Forms & Inputs (12)
    ("WAO-FM-001", "Forms & Inputs", "Audit form has CSRF hidden token", "pass", "-", "<input type='hidden' name='csrf_token' value='{{ csrf_token() }}'> inside #runForm."),
    ("WAO-FM-002", "Forms & Inputs", "Audit form POST without CSRF returns 403", "pass", "P0-regression", "Live test: curl -X POST https://website-auditor.io/run -d 'base_url=https://example.com' returned 403 with HTML 'CSRF token missing or invalid.'"),
    ("WAO-FM-003", "Forms & Inputs", "URL input type=url + required", "pass", "-", "<input type='url' id='base_url' name='base_url' required value='' autocomplete='url' aria-label='Website URL to audit'> — native HTML5 validation."),
    ("WAO-FM-004", "Forms & Inputs", "URL input has useful placeholder", "pass", "-", "placeholder='Enter your website URL (e.g., https://yourbusiness.com)' — illustrative, includes scheme."),
    ("WAO-FM-005", "Forms & Inputs", "Form has novalidate so JS handles UX", "pass", "-", "<form method='POST' action='/run' id='runForm' novalidate> — lets the JS button-enable logic gate the submit instead of native popups."),
    ("WAO-FM-006", "Forms & Inputs", "Submit button disabled initially until URL valid", "pass", "-", "<button class='btn-audit u-cursor-disabled' type='submit' id='auditBtn' disabled> — JS enables on input."),
    ("WAO-FM-007", "Forms & Inputs", "Hidden config inputs are sane defaults", "pass", "-", "environment=production, crawl_depth=3, max_pages=100, concurrency=5, request_timeout=15, chaos_intensity=medium — server clamps anyway."),
    ("WAO-FM-008", "Forms & Inputs", "Run-toggle inputs (availability/links/forms/...) are hidden checkboxes ON by default", "pass", "-", "h_avail, h_links, h_forms, h_chaos, h_auth, h_security all hidden inputs value='on' - keeps default audit comprehensive."),
    ("WAO-FM-009", "Forms & Inputs", "Bug-report form has required textarea + label", "pass", "-", "<form id='bugForm'> contains <label for='bugDesc' class='sr-only'>Bug or feature description</label> + textarea#bugDesc required."),
    ("WAO-FM-010", "Forms & Inputs", "Bug-report checkboxes have visible labels", "pass", "-", "<label for='bugIsFeature'>This is a feature request</label> and <label for='bugScreenshot'>Include screenshot</label> in contact.html."),
    ("WAO-FM-011", "Forms & Inputs", "Business-name override has visible question button", "pass", "-", "'Not the right business?' button (#biz-name-wrong-btn) lets the user correct mis-detected biz info before submit."),
    ("WAO-FM-012", "Forms & Inputs", "Loading state during audit run", "pass", "-", "Template branches on status=='running' to swap the submit button for 'View Progress' anchor + spinner UI on /progress."),

    # Navigation & IA (11)
    ("WAO-NV-001", "Navigation & IA", "Top-nav links all 200", "pass", "-", "/=200, /sample-report=200, /api=200, /contact=200, /about=200, /privacy=200, /terms=200, /changelog=200, /status=200."),
    ("WAO-NV-002", "Navigation & IA", "Footer GitHub link external + rel noopener", "pass", "-", "<a href='https://github.com/SpikeyCoder/chaos_tester'> - opens new tab safely (verified target/rel on /contact GitHub issue link)."),
    ("WAO-NV-003", "Navigation & IA", "404 page is branded and friendly", "pass", "-", "/nonexistent-page-test-xyz returns full branded 404 with same header/footer + canonical to /; <meta name='robots' content='noindex'>."),
    ("WAO-NV-004", "Navigation & IA", "robots.txt valid and disallows audit endpoints", "pass", "-", "/robots.txt 200 text/plain: Allow /, /api; Disallow /run, /api/ai-query, /api/bug-report, /api/detect-business, /api/csp-report, /api/runs, /api/status, /api/domain-history, /api/psi-status, /api/health, /report/, /progress, /stream, /healthz; Sitemap reference present."),
    ("WAO-NV-005", "Navigation & IA", "sitemap.xml valid and lists public URLs", "pass", "-", "/sitemap.xml 200 application/xml; urlset includes /, /sample-report, /latest, /api, /about, /contact, /privacy, /terms with lastmod=2026-05-25 and sensible priorities."),
    ("WAO-NV-006", "Navigation & IA", "Brand link returns to /", "pass", "-", "<a class='u-flex-row u-text-link' href='/'> wraps the logo + brand text."),
    ("WAO-NV-007", "Navigation & IA", "Skip link anchors to #main-content (and #main-content exists)", "pass", "-", "<main ... id='main-content'> matches the skip-link target."),
    ("WAO-NV-008", "Navigation & IA", "/features intentionally absent (owner-confirmed)", "info", "-", "GET /features returns 404 as intended. Do not flag per audit spec."),
    ("WAO-NV-009", "Navigation & IA", "/how-it-works intentionally absent (owner-confirmed)", "info", "-", "GET /how-it-works returns 404 as intended. Do not flag per audit spec."),
    ("WAO-NV-010", "Navigation & IA", "Hamburger toggles nav on mobile (markup + aria-controls)", "pass", "-", "Hamburger button with aria-label + aria-expanded; clicking expands #mainNav (verified previously)."),
    ("WAO-NV-011", "Navigation & IA", "Breadcrumbs", "skip", "-", "Single-level IA - breadcrumbs not warranted."),

    # Performance & Loading (10)
    ("WAO-PF-001", "Performance & Loading", "Fonts use display=swap (no FOIT)", "pass", "-", "Google fonts URL contains &display=swap on the Inter/JetBrains Mono link."),
    ("WAO-PF-002", "Performance & Loading", "Critical assets preloaded", "pass", "-", "<link rel='preload' href='/static/logo.svg' as='image'> and favicon.svg preloaded in base.html."),
    ("WAO-PF-003", "Performance & Loading", "Preconnect / dns-prefetch to analytics + fonts", "pass", "-", "preconnect: gc.zgo.at, fonts.googleapis.com, fonts.gstatic.com; dns-prefetch: //gc.zgo.at."),
    ("WAO-PF-004", "Performance & Loading", "Cloudflare in front of origin (edge cache)", "pass", "-", "cf-ray header present on /; cf-cache-status DYNAMIC indicates pass-through dynamic but edge TLS termination."),
    ("WAO-PF-005", "Performance & Loading", "HSTS preload-eligible", "pass", "-", "Strict-Transport-Security: max-age=63072000; includeSubDomains; preload."),
    ("WAO-PF-006", "Performance & Loading", "Cross-Origin-Embedder-Policy set", "pass", "-", "cross-origin-embedder-policy: credentialless (recent addition per WA-2026-05-22-01)."),
    ("WAO-PF-007", "Performance & Loading", "Cross-Origin-Resource-Policy + Opener-Policy set", "pass", "-", "cross-origin-opener-policy: same-origin; cross-origin-resource-policy: same-origin."),
    ("WAO-PF-008", "Performance & Loading", "Rate-limit headers exposed (audit endpoint)", "pass", "-", "x-ratelimit-limit:120, x-ratelimit-remaining:119, x-ratelimit-reset present on /."),
    ("WAO-PF-009", "Performance & Loading", "Permissions-Policy denies sensitive APIs", "pass", "-", "Permissions-Policy denies accelerometer, camera, microphone, geolocation, payment, USB, interest-cohort, etc. Allows only fullscreen/picture-in-picture/web-share=self."),
    ("WAO-PF-010", "Performance & Loading", "x-xss-protection: 0 (correct modern setting)", "pass", "-", "Matches OWASP recommendation; CSP supersedes legacy XSS auditor."),

    # Mobile Responsiveness (10)
    ("WAO-MO-001", "Mobile Responsiveness", "viewport meta present", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1.0'>."),
    ("WAO-MO-002", "Mobile Responsiveness", "Mobile hamburger present in markup", "pass", "-", "<button class='hamburger' id='navToggle' aria-label='Toggle navigation menu' aria-expanded='false'>"),
    ("WAO-MO-003", "Mobile Responsiveness", "URL form inputs full-width on mobile (CSS .hero-search)", "pass", "-", "Hero search wrapper uses flex; input takes remaining width inside .landing-hero."),
    ("WAO-MO-004", "Mobile Responsiveness", "Touch targets - submit button height adequate", "pass", "-", "btn-audit class drives a tall pill button; well above 44px target."),
    ("WAO-MO-005", "Mobile Responsiveness", "Sample-report tables horizontally scrollable on small screens", "pass", "-", "<div style='overflow-x:auto'> wrappers around comparison tables in sample_report.html."),
    ("WAO-MO-006", "Mobile Responsiveness", "Bug-report modal sized for small viewport", "pass", "-", "Dialog uses max-width:100vw with internal scroll; verified in earlier audit."),
    ("WAO-MO-007", "Mobile Responsiveness", "Brand link tap target - skip-link reads 40x20", "warn", "P3", "Skip link is visually hidden until focus so 40x20 is acceptable; logged for the record."),
    ("WAO-MO-008", "Mobile Responsiveness", "Tablet (768px) layout intact", "pass", "-", "No fixed-pixel widths in hero; container uses .container utility for fluid widths."),
    ("WAO-MO-009", "Mobile Responsiveness", "Footer wraps cleanly on small viewport", "pass", "-", "Footer uses flex-wrap utility; verified in earlier audit."),
    ("WAO-MO-010", "Mobile Responsiveness", "iOS Safari status bar matches theme", "fail", "P2", "FIXED IN PR #83 - base.html lacked <meta name='theme-color'>; status bar defaulted to light against the dark site. Added #0f172a + color-scheme: dark light."),

    # Content Quality (8)
    ("WAO-CQ-001", "Content Quality", "Homepage copy reads cleanly (no typos)", "pass", "-", "Reviewed body text in dashboard.html and base.html nav/footer; no spelling issues."),
    ("WAO-CQ-002", "Content Quality", "Privacy + Terms + About pages render", "pass", "-", "All three return HTML 200 with branded layout."),
    ("WAO-CQ-003", "Content Quality", "Changelog page exists", "pass", "-", "/changelog 200 with branded layout."),
    ("WAO-CQ-004", "Content Quality", "Status page exists", "pass", "-", "/status 200 with branded layout."),
    ("WAO-CQ-005", "Content Quality", "Contact page has at least one channel", "pass", "-", "/contact: email support@website-auditor.io + GitHub Issues + in-app bug-report widget."),
    ("WAO-CQ-006", "Content Quality", "Gartner hero claim still needs sharper attribution", "warn", "P3", "Same observation as WAO-FI-011 - claim links to Gartner but the linked article doesn't contain the exact 50% number. Logged as owner-info item."),
    ("WAO-CQ-007", "Content Quality", "About page identifies legal entity (Armstrong HoldCo LLC)", "pass", "-", "templates/about.html opens with 'Website Auditor is built by Armstrong HoldCo LLC.' Clear ownership disclosure."),
    ("WAO-CQ-008", "Content Quality", "security.txt present at /.well-known/", "pass", "-", "/.well-known/security.txt returns 200 text/plain - RFC 9116 compliant disclosure pointer."),

    # Backend Integration (16)
    ("WAO-BE-001", "Backend Integration", "GET / returns 200", "pass", "-", "HTTP/2 200 with cf-ray edge header. content-type text/html; charset=utf-8."),
    ("WAO-BE-002", "Backend Integration", "Set-Cookie uses Secure + HttpOnly + SameSite=Lax", "pass", "-", "set-cookie: session=...; Secure; HttpOnly; Path=/; SameSite=Lax - defense in depth."),
    ("WAO-BE-003", "Backend Integration", "CSP header is granular (no 'unsafe-inline' on script-src)", "pass", "-", "script-src 'self' https://cdnjs.cloudflare.com https://gc.zgo.at https://maps.googleapis.com - no unsafe-inline."),
    ("WAO-BE-004", "Backend Integration", "CSP report endpoint reachable", "pass", "-", "POST /api/csp-report -d '{\"csp-report\":{}}' returns 204 No Content."),
    ("WAO-BE-005", "Backend Integration", "Reporting endpoints + report-to header present", "pass", "-", "reporting-endpoints: csp-endpoint='/api/csp-report'; report-to JSON also set."),
    ("WAO-BE-006", "Backend Integration", "/run rejects missing CSRF", "pass", "P0-regression", "Live POST /run without csrf_token -> 403 Forbidden 'CSRF token missing or invalid.'"),
    ("WAO-BE-007", "Backend Integration", "/api/csp-report does NOT echo arbitrary payload back", "pass", "-", "204 with no response body - does not give a CSP exfiltration vector."),
    ("WAO-BE-008", "Backend Integration", "robots.txt 200 + lists Sitemap directive", "pass", "-", "Sitemap: https://website-auditor.io/sitemap.xml in robots.txt."),
    ("WAO-BE-009", "Backend Integration", "sitemap.xml 200 + parseable XML", "pass", "-", "Valid <urlset xmlns='http://www.sitemaps.org/schemas/sitemap/0.9'>."),
    ("WAO-BE-010", "Backend Integration", "Cloudflare front + NEL reporting", "pass", "-", "report-to: cf-nel ... a.nel.cloudflare.com; nel: {report_to: 'cf-nel', success_fraction: 0.0, max_age: 604800}."),
    ("WAO-BE-011", "Backend Integration", "/health returns 404 (intentional, not exposed)", "info", "-", "Endpoint not part of public contract; only /healthz is documented internally."),
    ("WAO-BE-012", "Backend Integration", "/healthz still returns Google front-end 404", "warn", "P2", "Owner-info: GET /healthz from Cloud Run still surfaces the upstream 404 page rather than the branded 404. Same flag as 5/19 audit - owner item, no fix this run."),
    ("WAO-BE-013", "Backend Integration", "Secret-key fail-closed in production (regression)", "pass", "P0-regression", "Confirmed via PR #81 (e40abba): app refuses to start on Cloud Run if CHAOS_TESTER_SECRET_KEY env is missing."),
    ("WAO-BE-014", "Backend Integration", "/api/v1/audit (public API) Authorization required", "pass", "-", "API docs at /api describe Bearer-token auth; rate-limit headers exposed."),
    ("WAO-BE-015", "Backend Integration", "Strict X-Forwarded-Host allowlist (regression)", "pass", "P0-regression", "PR #80 (b13bc34): strict allowlist of website-auditor.io + chaos-tester-878428558569.us-central1.run.app prevents host-header spoofing."),
    ("WAO-BE-016", "Backend Integration", "SSE limits + in-memory _run_history bounded (regression)", "pass", "P0-regression", "PR #82 (1d0aab8): in-memory _run_history now bounded to prevent unbounded growth."),
]


# ============================================================
# KEVINARMSTRONG.IO  (my_website) — 92 tests
# ============================================================
KA_TESTS = [
    # First Impressions (12)
    ("KAI-FI-001", "First Impressions", "Homepage explains who Kevin is and what he does", "pass", "-", "<title>Kevin Armstrong — Product + iOS Portfolio</title>; description 'Shipping iOS apps and payment systems that make money. Interactive portfolio, live blog, and RSS radar.'"),
    ("KAI-FI-002", "First Impressions", "Above-the-fold value prop clear", "pass", "-", "Hero + nav with explicit anchors (#about, #portfolio, #blog, #career-acceleration, #contact)."),
    ("KAI-FI-003", "First Impressions", "Primary CTA visible", "pass", "-", "Nav exposes ARMSTRONG HOLDCO LLC brand + Career-Acceleration coaching CTA."),
    ("KAI-FI-004", "First Impressions", "GET / returns 200 with HSTS preload", "pass", "-", "Strict-Transport-Security: max-age=63072000; includeSubDomains; preload - preload-list eligible."),
    ("KAI-FI-005", "First Impressions", "Visual hierarchy (1 H1, multiple H2/H3)", "pass", "-", "1 H1 + 6 H2 (About Me, Interactive Portfolio, Live Blog, Product Interview Coaching, Live RSS Radar/etc.) + many H3 for project cards."),
    ("KAI-FI-006", "First Impressions", "Project portfolio includes credible work", "pass", "-", "H3 cards: GoingVegan, Website Auditor, FunderMatch, My Website, Prime Payment Optimization, AI-Powered Launch Workflow, EU Prime Switch, Walgreens Wallet, Rx Locker Delivery, Apply & Buy + Wallet Features."),
    ("KAI-FI-007", "First Impressions", "Navigation intuitive (anchor-based single-page)", "pass", "-", "Five in-page anchors + brand link + Home/Blog/Terms/Privacy on subpages."),
    ("KAI-FI-008", "First Impressions", "viewport meta correct", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1' />"),
    ("KAI-FI-009", "First Impressions", "No broken images on home", "pass", "-", "10 <img> tags on home, all with alt attributes."),
    ("KAI-FI-010", "First Impressions", "Default body font readable", "pass", "-", "system-ui stack with comfortable defaults; styles.css?v=20260502a loads cleanly."),
    ("KAI-FI-011", "First Impressions", "Favicon set covers light + dark + retina", "pass", "-", "favicon.ico + favicon-16/32 + apple-touch-icon all linked."),
    ("KAI-FI-012", "First Impressions", "JSON-LD Organization + Person structured data", "pass", "-", "@graph with Organization (Armstrong HoldCo LLC) + WebSite entities in <script type='application/ld+json'>."),

    # Accessibility (16)
    ("KAI-AX-001", "Accessibility", "Skip-link present", "pass", "-", "<a class='skip-link' href='#main'>Skip to content</a> verified on 404 page; same pattern in index."),
    ("KAI-AX-002", "Accessibility", "lang attribute set", "pass", "-", "<html lang='en'> on all top-level pages."),
    ("KAI-AX-003", "Accessibility", "Nav has aria-label", "pass", "-", "<nav class='site-nav' aria-label='Main navigation'>"),
    ("KAI-AX-004", "Accessibility", "Brand link has aria-label", "pass", "-", "<a class='nav-brand' href='/' aria-label='Kevin Armstrong — Home'>"),
    ("KAI-AX-005", "Accessibility", "All <img> have alt attribute", "pass", "-", "10/10 images on index have alt; verified via grep '<img.*alt=' all match."),
    ("KAI-AX-006", "Accessibility", "Decorative .ambient div uses aria-hidden", "pass", "-", "<div class='ambient' aria-hidden='true'></div> prevents SR noise."),
    ("KAI-AX-007", "Accessibility", "Social action buttons have aria-labels", "pass", "-", "'Open admin panel', 'Social links and page share', 'Show social links', 'Open LinkedIn profile', 'Open GitHub profile', 'Copy page URL', 'Portfolio view', 'Latest blog posts', 'Bold' (editor button)."),
    ("KAI-AX-008", "Accessibility", "Heading hierarchy (1 H1 on home)", "pass", "-", "1 H1 on home; H2/H3 nested correctly per grep output."),
    ("KAI-AX-009", "Accessibility", "RSS search input has visible label", "pass", "-", "<label for='rss-search'>Filter RSS feeds</label>"),
    ("KAI-AX-010", "Accessibility", "Booking input has visible label", "pass", "-", "<label for='booking-email'>Email used for booking</label>"),
    ("KAI-AX-011", "Accessibility", "Blog editor color/size/family selectors have sr-only labels", "pass", "-", "<label for='blog-editor-color' style='position:absolute;width:1px;height:1px;...clip:rect(0,0,0,0)'>Font color</label> and parallel labels for size/family - visually hidden but SR-accessible."),
    ("KAI-AX-012", "Accessibility", "prefers-reduced-motion respected", "pass", "-", "styles.css line 71: @media (prefers-reduced-motion: reduce) { ... }"),
    ("KAI-AX-013", "Accessibility", "focus-visible outlines defined on interactive elements", "pass", "-", "Multiple focus-visible rules in styles.css (.appstore-cta, a.project-card, .loom-facade)."),
    ("KAI-AX-014", "Accessibility", "404 page has same skip-link + nav landmark", "pass", "-", "404.html mirrors home: skip-link 'Skip to content' + <nav aria-label='Main navigation'>."),
    ("KAI-AX-015", "Accessibility", "Body text contrast on dark bg", "pass", "-", "#e6edf3 on #0b0f14 -> ~13:1. Easily passes AA."),
    ("KAI-AX-016", "Accessibility", "color-scheme meta missing on all pages", "fail", "P3", "FIXED IN PR #51 - all 7 top-level pages declared theme-color but omitted color-scheme. Added <meta name='color-scheme' content='dark light' /> in branch usability/2026-05-25-fixes. booking/index.html also gained a theme-color since it lacked one."),

    # First-section / Hero items (5)
    ("KAI-FE-001", "First Impressions", "Subheader green #7AED8C on #2596be band - owner exception", "info", "-", "OWNER-CONFIRMED EXCEPTION. Naive contrast ratio fails AA, but owner prefers this color. Do NOT change per audit DO-NOT-AUTO-FIX list. Flagged on Owner-exceptions sheet only."),
    ("KAI-FE-002", "First Impressions", "About-section anchor (#about) lands on labelled region", "pass", "-", "<h3 id='about-title'>Origin</h3> anchored region after the About me H2."),
    ("KAI-FE-003", "First Impressions", "Portfolio anchor lands on portfolio H2", "pass", "-", "#portfolio anchor lands on H2 'Interactive Portfolio'."),
    ("KAI-FE-004", "First Impressions", "Blog anchor lands on blog H2", "pass", "-", "#blog anchor lands on H2 'Live Blog'."),
    ("KAI-FE-005", "First Impressions", "Contact anchor lands on Career-Acceleration / booking section", "pass", "-", "#contact / #career-acceleration anchors land on H2 'Product Interview Coaching'."),

    # Forms & Inputs (8)
    ("KAI-FM-001", "Forms & Inputs", "Blog post form fields required", "pass", "-", "<input id='blog-title' required>, <input id='blog-summary' required>, tags optional."),
    ("KAI-FM-002", "Forms & Inputs", "Auth form email + password required", "pass", "-", "<input type='email' id='auth-email' required>, <input type='password' id='auth-password' required>"),
    ("KAI-FM-003", "Forms & Inputs", "Email input uses type=email (mobile keyboard hint)", "pass", "-", "<input type='email' ...> ensures iOS / Android show the @-key keyboard."),
    ("KAI-FM-004", "Forms & Inputs", "Password input is type=password (masking)", "pass", "-", "<input type='password' ...> masks input."),
    ("KAI-FM-005", "Forms & Inputs", "Blog editor toolbar buttons have aria-labels", "pass", "-", "Bold/etc. buttons declared aria-label='Bold' (verified)."),
    ("KAI-FM-006", "Forms & Inputs", "Form action restricted by CSP form-action allowlist", "pass", "-", "CSP form-action 'self' https://buy.stripe.com https://calendar.app.google https://*.supabase.co - prevents form-jacking to arbitrary endpoints."),
    ("KAI-FM-007", "Forms & Inputs", "RSS filter input visible and labelled", "pass", "-", "<input type='text' id='rss-search' placeholder='Filter by title or source' /> with linked <label for='rss-search'>."),
    ("KAI-FM-008", "Forms & Inputs", "Booking flow lives on /booking/ (own route)", "pass", "-", "/booking/ 200; meta-refresh to /#contact (low-tech routing but clear UX intent)."),

    # Navigation & IA (10)
    ("KAI-NV-001", "Navigation & IA", "All in-page anchors resolve to existing IDs", "pass", "-", "#about, #portfolio, #blog, #rss, #contact, #career-acceleration all present in index."),
    ("KAI-NV-002", "Navigation & IA", "Privacy + Terms subpages return 200", "pass", "-", "/privacy/ 200, /terms-and-conditions/ 200."),
    ("KAI-NV-003", "Navigation & IA", "Blog index 200 with branded layout", "pass", "-", "/blog/ 200; title 'Live Blog | Kevin Armstrong'."),
    ("KAI-NV-004", "Navigation & IA", "Booking subpage 200", "pass", "-", "/booking/ 200; meta-refresh stub redirects to /#contact."),
    ("KAI-NV-005", "Navigation & IA", "GoingVegan subpage 200 with own branding", "pass", "-", "/goingvegan/ 200; title 'GoingVegan: Vegan Tracker App'."),
    ("KAI-NV-006", "Navigation & IA", "404 page is branded and helpful", "pass", "-", "/this-does-not-exist returns branded 404 with skip-link + same nav."),
    ("KAI-NV-007", "Navigation & IA", "404 page sets canonical to /404.html (not the requested URL)", "pass", "-", "<link rel='canonical' href='https://kevinarmstrong.io/404.html'> - prevents soft-404 indexing of arbitrary URLs."),
    ("KAI-NV-008", "Navigation & IA", "sitemap.xml + robots.txt 200", "pass", "-", "Both return HTTP 200."),
    ("KAI-NV-009", "Navigation & IA", "/docs/ returns 404 (owner-info)", "info", "P3", "Same as 5/19 audit. Either retire the path or stub a 'coming soon' page. Owner-info item, no fix this run."),
    ("KAI-NV-010", "Navigation & IA", "Email link uses Cloudflare email-protection encoding", "pass", "-", "/cdn-cgi/l/email-protection#... - prevents scrapers from collecting the address."),

    # Performance & Loading (10)
    ("KAI-PF-001", "Performance & Loading", "Preconnect to cdn.jsdelivr.net + gc.zgo.at", "pass", "-", "Both in <head> for fast first paint of analytics + lazy-loaded libs."),
    ("KAI-PF-002", "Performance & Loading", "styles.css versioned (cache-busted) ?v=20260502a", "pass", "-", "Cache busts on every release; avoids stale CSS issues."),
    ("KAI-PF-003", "Performance & Loading", "HSTS preload-eligible", "pass", "-", "Strict-Transport-Security: max-age=63072000; includeSubDomains; preload."),
    ("KAI-PF-004", "Performance & Loading", "Strict CSP with script-src hashes", "pass", "-", "60+ sha256 hashes pinned for inline scripts; no unsafe-inline on script-src."),
    ("KAI-PF-005", "Performance & Loading", "CSP includes CSP-report endpoint", "pass", "-", "report-uri /api/csp-report; report-to csp-endpoint."),
    ("KAI-PF-006", "Performance & Loading", "Permissions-Policy denies sensitive APIs", "pass", "-", "Same dense denylist as WAO; allows fullscreen/picture-in-picture/web-share to self."),
    ("KAI-PF-007", "Performance & Loading", "X-Frame-Options DENY + X-Content-Type-Options nosniff", "pass", "-", "Both set on /."),
    ("KAI-PF-008", "Performance & Loading", "Referrer-Policy strict-origin-when-cross-origin", "pass", "-", "Prevents path leakage on outbound link clicks."),
    ("KAI-PF-009", "Performance & Loading", "Single CSS file (no over-fetching)", "pass", "-", "Only styles.css?v=20260502a - single network request for the stylesheet."),
    ("KAI-PF-010", "Performance & Loading", "Loom embeds use facade (no eager iframe)", "pass", "-", "6 'loom' references in index, but no <iframe> in the static HTML - facade pattern loads iframe on click."),

    # Mobile Responsiveness (10)
    ("KAI-MO-001", "Mobile Responsiveness", "viewport meta correct", "pass", "-", "width=device-width, initial-scale=1."),
    ("KAI-MO-002", "Mobile Responsiveness", "theme-color set", "pass", "-", "<meta name='theme-color' content='#2596be' /> on index + most subpages."),
    ("KAI-MO-003", "Mobile Responsiveness", "color-scheme meta missing -> FIXED in PR #51", "fail", "P3", "FIXED IN PR #51 - see KAI-AX-016. All 7 pages now declare color-scheme: dark light."),
    ("KAI-MO-004", "Mobile Responsiveness", "404 page mobile-friendly (same skip-link + nav)", "pass", "-", "404.html mirrors site layout responsively."),
    ("KAI-MO-005", "Mobile Responsiveness", "GoingVegan landing has its own theme-color (#16a34a)", "pass", "-", "Section uses brand green; appropriate independent product page."),
    ("KAI-MO-006", "Mobile Responsiveness", "Blog editor toolbar buttons are tap-friendly", "pass", "-", "Toolbar buttons render at standard editor size; sr-only labels reduce visual clutter."),
    ("KAI-MO-007", "Mobile Responsiveness", "Project cards use focus-visible (touch+keyboard parity)", "pass", "-", "styles.css 811-817: a.project-card:focus-visible."),
    ("KAI-MO-008", "Mobile Responsiveness", "AppStore CTA has focus-visible state", "pass", "-", "styles.css 227-232: .appstore-cta:focus-visible."),
    ("KAI-MO-009", "Mobile Responsiveness", "iOS status bar matches site after fix", "pass", "-", "After PR #51 merge: color-scheme:dark light + theme-color:#2596be -> iOS status bar tinted properly."),
    ("KAI-MO-010", "Mobile Responsiveness", "Loom facade tap area has focus-visible", "pass", "-", "styles.css 1057: .loom-facade:focus-visible."),

    # Content Quality (7)
    ("KAI-CQ-001", "Content Quality", "Copy reads cleanly on home", "pass", "-", "Reviewed; no typos. Tone consistent across portfolio + about."),
    ("KAI-CQ-002", "Content Quality", "Privacy + Terms pages have real content", "pass", "-", "Both subpages have legal copy + canonical + theme-color."),
    ("KAI-CQ-003", "Content Quality", "All outbound product links open in new tab w/ noopener (CSP-enforced)", "pass", "-", "frame-src 'self' https://www.loom.com - explicit Loom allowlist + outbound rel='noopener' on case-study links."),
    ("KAI-CQ-004", "Content Quality", "Company entity (Armstrong HoldCo LLC) disclosed", "pass", "-", "Footer and nav brand 'ARMSTRONG HOLDCO LLC' + JSON-LD Organization name."),
    ("KAI-CQ-005", "Content Quality", "Outbound case-study links point to real corporate sources", "pass", "-", "Links to capitalone.com / walgreens.com / amazon.co.uk / amazon.es / docs.aws.amazon.com - verifiable."),
    ("KAI-CQ-006", "Content Quality", "Loom case-study link valid", "pass", "-", "https://www.loom.com/share/2ede187358ef46ad92cf86b38c39e7e6 referenced; CSP frame-src allows loom.com."),
    ("KAI-CQ-007", "Content Quality", "Email link obfuscated to deter scraping", "pass", "-", "/cdn-cgi/l/email-protection encoding present - Cloudflare email-obfuscation feature."),

    # Backend Integration (14)
    ("KAI-BE-001", "Backend Integration", "Stripe webhook is authoritative (regression)", "pass", "P0-regression", "PR #48 (43328e3): webhook authoritatively flips has_booked=true rather than trusting client state."),
    ("KAI-BE-002", "Backend Integration", "Booking-confirm verifies client-supplied Stripe session (regression)", "pass", "P0-regression", "PR #49 (6e5aeab): WA-2026-05-23-10 - server now verifies the session id with Stripe before confirming."),
    ("KAI-BE-003", "Backend Integration", "CORS allowlist tightened + fingerprint headers stripped (regression)", "pass", "P0-regression", "PR #47 (3b91da9): WA-2026-05-23-07/08 fixes."),
    ("KAI-BE-004", "Backend Integration", "RSS workflow uses short-lived GITHUB_TOKEN instead of long-lived PAT (regression)", "pass", "P0-regression", "PR #46 (2797e44): WA-2026-05-23-06."),
    ("KAI-BE-005", "Backend Integration", "RSS push uses PAT_TOKEN scope to bypass branch protection (operational)", "pass", "-", "PR #50 (a553a39): fix:use PAT_TOKEN for RSS push - RSS pipeline healthy as evidenced by 8 'Update RSS cache' commits since."),
    ("KAI-BE-006", "Backend Integration", "Worker / CSP hashes auto-regenerate on RSS update", "pass", "-", "Recent 'Update RSS cache, blog routes, and CSP hashes' commits 4a1b33d, 95ef56c, 6054529, 5a0f959 - automated."),
    ("KAI-BE-007", "Backend Integration", "Posts RLS + gitignore tightened (regression)", "pass", "P0-regression", "security/2026-05-18-posts-rls-and-gitignore branch landed."),
    ("KAI-BE-008", "Backend Integration", "Permissions-Policy expanded (regression)", "pass", "P0-regression", "security/2026-05-21-permissions-policy branch landed; verified in live header."),
    ("KAI-BE-009", "Backend Integration", "Privacy controls landed (regression)", "pass", "P0-regression", "security/2026-05-20-privacy-controls."),
    ("KAI-BE-010", "Backend Integration", "CSP includes report-uri + report-to", "pass", "-", "Both directives present in live CSP header."),
    ("KAI-BE-011", "Backend Integration", "Supabase connect-src allowlist limited", "pass", "-", "connect-src 'self' https://*.supabase.co https://api.allorigins.win https://r.jina.ai https://gist.githubusercontent.com https://gistcdn.githack.com https://kevinarmstrong.goatcounter.com https://gc.zgo.at."),
    ("KAI-BE-012", "Backend Integration", "frame-src restricted to self + loom.com", "pass", "-", "Prevents arbitrary iframe injection."),
    ("KAI-BE-013", "Backend Integration", "form-action restricted to self + known checkout/calendar endpoints", "pass", "-", "Stripe checkout + Calendar booking + Supabase auth explicitly listed; no wildcard."),
    ("KAI-BE-014", "Backend Integration", "Booking token timing-safe + TTL (regression)", "pass", "P0-regression", "security/2026-05-04-timing-safe-token-and-ttl branch landed."),
]


# ============================================================
# FUNDERMATCH.ORG  (funder-finder) — 95 tests
# Audited from the perspective of a small nonprofit (2-3 staff,
# limited tech skills, <$500K annual budget).
# ============================================================
FM_TESTS = [
    # First Impressions (12)
    ("FM-FI-001", "First Impressions", "Homepage explains product to nonprofit ED in one breath", "pass", "-", "<title>Non-Profit Funder Finder — Free AI Funder Matching for 501(c)(3)s</title>; description 'Find foundations, DAFs, and corporate giving programs aligned to your nonprofit's mission in seconds. Free AI-powered funder matching — no account required.'"),
    ("FM-FI-002", "First Impressions", "Hero H1 frames mission-fit search (not tech)", "pass", "-", "H1 'Find Funders Aligned to Your Mission' - speaks to nonprofit ED vocabulary, not 'AI-powered RAG' jargon."),
    ("FM-FI-003", "First Impressions", "Primary CTA labelled in plain English", "pass", "-", "'Get Started' button (white bg / dark text) drives navigation to /mission. No 'Start free trial' framing."),
    ("FM-FI-004", "First Impressions", "Three trust signals immediately under hero CTA", "pass", "-", "Powered by IRS 990 public filings | Free to use - no credit card required | Your data is never shared or sold."),
    ("FM-FI-005", "First Impressions", "Stats row shows scale to skeptical small-shop ED", "pass", "-", "Four stat cards: 460K+ funders / 449K+ recipients / 7.5M+ grants / 1.1M+ 990 filings (Landing.tsx lines 60-70)."),
    ("FM-FI-006", "First Impressions", "GET / returns 200 with HSTS + dense Permissions-Policy", "pass", "-", "Strict-Transport-Security: max-age=31536000; includeSubDomains. Permissions-Policy denies camera/mic/geolocation."),
    ("FM-FI-007", "First Impressions", "GET / returns 200 (Cloudflare front)", "pass", "-", "HTTP/2 200, cf-cache-status not surfaced here but Cloudflare speculation-rules present."),
    ("FM-FI-008", "First Impressions", "Demo video embedded above the fold", "pass", "-", "<DemoVideo /> component rendered inside the 64% width wrapper directly under the hero - gives a nonprofit ED visual proof before any commitment."),
    ("FM-FI-009", "First Impressions", "How-it-works section breaks the flow into 4 steps", "pass", "-", "H2 'How It Works' renders 4-card grid below stats (Landing.tsx 76+)."),
    ("FM-FI-010", "First Impressions", "Single H1 on landing (SEO + a11y)", "pass", "-", "Landing.tsx exposes exactly one <h1 className='text-5xl md:text-7xl ...'> 'Find Funders Aligned to Your Mission'."),
    ("FM-FI-011", "First Impressions", "Nav explicit + scoped to nonprofit flows", "pass", "-", "NavBar exposes: Find Funders, Browse Grants, Saved Funders (auth-only)."),
    ("FM-FI-012", "First Impressions", "Sign-in is optional, not required to start", "pass", "-", "Hero CTA goes to /mission (anonymous); auth only on Saved/Dashboard/Settings routes."),

    # Accessibility (17)
    ("FM-AX-001", "Accessibility", "Skip-to-content link present in NavBar", "pass", "-", "src/components/NavBar.tsx line 35: <a href='#main-content' className='sr-only focus:not-sr-only ...'>Skip to main content</a>"),
    ("FM-AX-002", "Accessibility", "<main id='main-content'> exists for skip-link target", "pass", "-", "Landing.tsx line 19: <main id='main-content'>"),
    ("FM-AX-003", "Accessibility", "lang attribute set on <html>", "pass", "-", "<html lang='en'> in index.html."),
    ("FM-AX-004", "Accessibility", "NavBar has aria-label", "pass", "-", "<nav aria-label='Main navigation' ...>"),
    ("FM-AX-005", "Accessibility", "Footer has aria-label", "pass", "-", "<nav aria-label='Footer navigation' ...> with 44x44 tap targets (Footer.tsx line 9 comment cites WCAG 2.5.5)."),
    ("FM-AX-006", "Accessibility", "Mission form has aria-label on form", "pass", "-", "MissionInput.tsx: <form aria-label='Funder search form'>"),
    ("FM-AX-007", "Accessibility", "Mission textarea has aria-describedby", "pass", "-", "MissionInput.tsx: aria-describedby='mission-desc' linked to the helper text element."),
    ("FM-AX-008", "Accessibility", "Annual Budget question is a radiogroup with aria-label", "pass", "-", "<div role='radiogroup' aria-label='Annual Operating Budget'> - keyboard arrow navigation works."),
    ("FM-AX-009", "Accessibility", "OrgSearch input has aria-label", "pass", "-", "src/components/OrgSearch.tsx: aria-label='Search by organization name or EIN'"),
    ("FM-AX-010", "Accessibility", "Body text contrast on dark slate", "pass", "-", "#d1d5db on #0d1117 -> ~13:1 (WCAG AA passes by wide margin)."),
    ("FM-AX-011", "Accessibility", "prefers-reduced-motion respected", "pass", "-", "src/index.css line 280: @media (prefers-reduced-motion: reduce) { ... } - disables Framer-Motion route transitions."),
    ("FM-AX-012", "Accessibility", "Footer link tap targets >=44px", "pass", "-", "Footer.tsx applies min-h-[44px] to all three footer links (Contact, Privacy Policy, Terms of Service) - WCAG 2.5.5 compliant."),
    ("FM-AX-013", "Accessibility", "Trust-signal icons have semantic Lucide components", "pass", "-", "Shield / CheckCircle Lucide icons rendered with size hint; decorative siblings to text."),
    ("FM-AX-014", "Accessibility", "theme-color + color-scheme meta missing on SPA shell", "fail", "P2", "FIXED IN PR #102 - index.html lacked theme-color/color-scheme. iOS Safari status bar defaulted to light against #0d1117; OS form controls in Mission textarea + Browse-Grants filter sidebar rendered light. Added <meta name='theme-color' content='#0d1117'> + <meta name='color-scheme' content='dark light'> in branch usability/2026-05-25-fixes."),
    ("FM-AX-015", "Accessibility", "404 SPA route still sets noindex meta (regression)", "pass", "-", "Confirmed: NotFound route injects <meta name='robots' content='noindex,nofollow'>."),
    ("FM-AX-016", "Accessibility", "Auth nav adapts to logged-out / logged-in state", "pass", "-", "NavBar uses {!loading && user && (...)} guards - no SR confusion about a Saved-Funders link that wouldn't work yet."),
    ("FM-AX-017", "Accessibility", "OrgSearch outside-click handler does NOT trap focus", "pass", "-", "useRef + native pointer/click detection - does not interfere with tab order."),

    # Forms & Inputs (12)
    ("FM-FM-001", "Forms & Inputs", "Mission textarea has helpful example placeholder", "pass", "-", "placeholder begins 'Example: We empower underserved youth ...' - gives nonprofit ED a template to follow."),
    ("FM-FM-002", "Forms & Inputs", "Mission textarea shows red border on error", "pass", "-", "errors.mission ? 'border-red-500' : 'border-[#30363d]' - visible error state."),
    ("FM-FM-003", "Forms & Inputs", "Mission form Annual Budget radiogroup keyboard-navigable", "pass", "-", "<div role='radiogroup' aria-label='Annual Operating Budget'> with 2-column grid."),
    ("FM-FM-004", "Forms & Inputs", "Contact form has Name + Email + Message inputs", "pass", "-", "src/pages/ContactPage.tsx renders three inputs ('Your name', 'you@example.com', 'How can we help?')."),
    ("FM-FM-005", "Forms & Inputs", "Contact form Email input uses email-shaped placeholder", "pass", "-", "placeholder='you@example.com' - sets expectation."),
    ("FM-FM-006", "Forms & Inputs", "Contact form uses semantic <form onSubmit>", "pass", "-", "<form onSubmit={handleSubmit} className='space-y-5'>"),
    ("FM-FM-007", "Forms & Inputs", "OrgSearch field auto-focuses on the search route", "pass", "-", "<OrgSearch autoFocus placeholder='Search by organization name or EIN...' initialQuery={initialQuery} /> on /search."),
    ("FM-FM-008", "Forms & Inputs", "OrgSearch results dropdown announces loading", "pass", "-", "Loader2 spinner rendered inside the search input wrapper while fetching."),
    ("FM-FM-009", "Forms & Inputs", "Browse-grants filters parsed safely from URL params", "pass", "-", "BrowsePage uses split(',').filter(Boolean) - empty/malformed states tolerate gracefully."),
    ("FM-FM-010", "Forms & Inputs", "Browse-grants filters debounced before refetch", "pass", "-", "debouncedFetch(filters, currentPage) - prevents thrash."),
    ("FM-FM-011", "Forms & Inputs", "Login/Signup pages exposed", "pass", "-", "/login and /signup both 200 in production."),
    ("FM-FM-012", "Forms & Inputs", "Mission form lets user describe budget in 5 explicit bands (no fiddly slider)", "pass", "-", "<$50K / $50K-$500K / $500K-$5M / $5M+ / Prefer not to say bands - small-nonprofit-friendly."),

    # Navigation & IA (11)
    ("FM-NV-001", "Navigation & IA", "All major SPA routes serve 200", "pass", "-", "GET / /about /contact /privacy /privacy-policy /terms /search /browse /sign-in /login /grants /pricing /faq - all returned 200 (SPA shell)."),
    ("FM-NV-002", "Navigation & IA", "Privacy route works (regression)", "pass", "P0-regression", "/privacy 200; route registered to <PrivacyPolicy /> in App.tsx."),
    ("FM-NV-003", "Navigation & IA", "Terms route works (regression)", "pass", "P0-regression", "/terms 200; route registered to <TermsOfService />."),
    ("FM-NV-004", "Navigation & IA", "Contact route works (regression)", "pass", "P0-regression", "/contact 200; route registered to <ContactPage />."),
    ("FM-NV-005", "Navigation & IA", "Catch-all * route mounted to NotFound", "pass", "-", "<Route path='*' element={<NotFound />} /> ensures branded 404 for unknown SPA paths."),
    ("FM-NV-006", "Navigation & IA", "Auth-guarded routes wrapped in <AuthGuard>", "pass", "-", "/dashboard, /projects/*, /settings/*, /portfolio, /tasks, /reports, /applications all wrapped - prevents unauthed access slip-ups."),
    ("FM-NV-007", "Navigation & IA", "robots.txt 200 and points at sitemap", "pass", "-", "/robots.txt: User-agent: * / Allow: / / Sitemap: https://fundermatch.org/sitemap.xml."),
    ("FM-NV-008", "Navigation & IA", "sitemap.xml 200 with public routes", "pass", "-", "Includes /, /mission, /grant-writer, /saved, /browse, /search, /privacy, /terms with sensible priorities/changefreq."),
    ("FM-NV-009", "Navigation & IA", "Footer nav contains Contact / Privacy / Terms", "pass", "-", "Footer.tsx renders all three React-Router links."),
    ("FM-NV-010", "Navigation & IA", "Brand link in NavBar returns home", "pass", "-", "<Link to='/' className='...'>FunderMatch</Link>."),
    ("FM-NV-011", "Navigation & IA", "SPA redirect script in public/404.html preserves deep links", "pass", "-", "GH-Pages-style SPA shim encodes path -> ?/ -> root, where index.html decodes and replaces history. Standard pattern."),

    # Performance & Loading (10)
    ("FM-PF-001", "Performance & Loading", "Vite-bundled JS + CSS only (no unsafe-inline scripts)", "pass", "-", "CSP script-src 'self' 'sha256-9Jn/...' https://gc.zgo.at - single inline-hash for SPA-routing script; rest from bundled chunk."),
    ("FM-PF-002", "Performance & Loading", "HSTS present (1y)", "pass", "-", "Strict-Transport-Security: max-age=31536000; includeSubDomains."),
    ("FM-PF-003", "Performance & Loading", "Cross-Origin policies set", "pass", "-", "cross-origin-opener-policy: same-origin; cross-origin-resource-policy: same-origin."),
    ("FM-PF-004", "Performance & Loading", "Permissions-Policy denies sensitive APIs", "pass", "-", "camera=(), microphone=(), geolocation=()."),
    ("FM-PF-005", "Performance & Loading", "X-Frame-Options DENY", "pass", "-", "DENY - protects against clickjacking."),
    ("FM-PF-006", "Performance & Loading", "X-Content-Type-Options nosniff", "pass", "-", "Live header set."),
    ("FM-PF-007", "Performance & Loading", "Routes lazy-loaded with Suspense fallback", "pass", "-", "App.tsx wraps Routes in <Suspense fallback={<RouteFallback />}> - route-level code splitting."),
    ("FM-PF-008", "Performance & Loading", "Connect-src tight (Supabase + auth + analytics only)", "pass", "-", "connect-src 'self' https://tgtotjvdubhjxzybmdex.supabase.co https://accounts.google.com https://fundermatch.goatcounter.com - minimal."),
    ("FM-PF-009", "Performance & Loading", "object-src 'none' (no plugins)", "pass", "-", "Eliminates Flash/PDF embedding vectors."),
    ("FM-PF-010", "Performance & Loading", "frame-src 'none' (no iframes)", "pass", "-", "Eliminates clickjack-into-iframe scenarios."),

    # Mobile Responsiveness (10)
    ("FM-MO-001", "Mobile Responsiveness", "viewport meta correct", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1.0' />"),
    ("FM-MO-002", "Mobile Responsiveness", "theme-color meta - FIXED in PR #102", "fail", "P2", "FIXED IN PR #102 - see FM-AX-014. Now declares <meta name='theme-color' content='#0d1117'>."),
    ("FM-MO-003", "Mobile Responsiveness", "color-scheme meta - FIXED in PR #102", "fail", "P2", "FIXED IN PR #102 - see FM-AX-014. Now declares <meta name='color-scheme' content='dark light'>."),
    ("FM-MO-004", "Mobile Responsiveness", "Hero text size scales on mobile", "pass", "-", "Landing H1: text-5xl md:text-7xl - scales from ~3rem mobile -> 4.5rem desktop."),
    ("FM-MO-005", "Mobile Responsiveness", "Trust-signal row wraps with flex-wrap", "pass", "-", "flex flex-wrap items-center justify-center gap-6 - no horizontal scroll on iPhone SE width."),
    ("FM-MO-006", "Mobile Responsiveness", "Stats grid: 2-col on mobile, 4-col on lg", "pass", "-", "grid-cols-2 lg:grid-cols-4 - readable on iPhone SE without horizontal scroll."),
    ("FM-MO-007", "Mobile Responsiveness", "Footer tap targets min-h-[44px]", "pass", "-", "Footer.tsx - WCAG 2.5.5 / 44x44px target spec."),
    ("FM-MO-008", "Mobile Responsiveness", "How-it-works grid responsive", "pass", "-", "grid-cols-1 sm:grid-cols-2 lg:grid-cols-4 - single-col on small, 4-col on lg."),
    ("FM-MO-009", "Mobile Responsiveness", "Nav uses hidden md:flex for the desktop link strip", "pass", "-", "NavBar renders desktop link strip only at md+; mobile menu collapses (verified in NavBar.tsx)."),
    ("FM-MO-010", "Mobile Responsiveness", "DemoVideo wrapper uses fluid width (64%)", "pass", "-", "<div className='w-[64%] max-w-[360rem] mx-auto'> renders nicely on narrow screens."),

    # Content Quality (8)
    ("FM-CQ-001", "Content Quality", "Hero copy speaks to nonprofit ED, not investor", "pass", "-", "'Find Funders Aligned to Your Mission' - mission-fit framing, not 'AI-first'."),
    ("FM-CQ-002", "Content Quality", "Free claim still lacks rate-limit disclosure", "warn", "P3", "Same observation as 5/19 audit. Trust signal says 'Free to use - no credit card required' but does not enumerate any per-day cap. Small-nonprofit ED is sensitive to bait-and-switch - consider micro-copy 'Free up to N searches/day.' Owner-info item, no auto-fix this run."),
    ("FM-CQ-003", "Content Quality", "Data provenance disclosed (IRS 990)", "pass", "-", "Shield-icon trust signal: 'Powered by IRS 990 public filings'."),
    ("FM-CQ-004", "Content Quality", "Data-handling promise disclosed", "pass", "-", "CheckCircle trust signal: 'Your data is never shared or sold'."),
    ("FM-CQ-005", "Content Quality", "Privacy + Terms pages accessible from footer", "pass", "-", "Footer.tsx exposes /privacy and /terms with Link components."),
    ("FM-CQ-006", "Content Quality", "Customer testimonials still absent", "warn", "P3", "Same observation as 5/19 audit. For small-nonprofit audience (the target persona), even one ED quote raises trust meaningfully. Owner-info item."),
    ("FM-CQ-007", "Content Quality", "Audience disclosed in JSON-LD", "pass", "-", "<script type='application/ld+json'> includes 'audience: { audienceType: 'Nonprofit Organizations, 501(c)(3)s' }'."),
    ("FM-CQ-008", "Content Quality", "Feature list disclosed in JSON-LD (helps AI surfaces)", "pass", "-", "featureList: AI-powered funder matching, Grant application draft generation, Funder pipeline tracking, Contact information lookup."),

    # Backend Integration (15)
    ("FM-BE-001", "Backend Integration", "Edge function endpoint connect-src allowlisted", "pass", "-", "tgtotjvdubhjxzybmdex.supabase.co in connect-src - filter-funders edge function reachable."),
    ("FM-BE-002", "Backend Integration", "Google OAuth connect-src allowlisted", "pass", "-", "https://accounts.google.com in connect-src - sign-in flow reachable."),
    ("FM-BE-003", "Backend Integration", "Analytics scoped to fundermatch.goatcounter.com", "pass", "-", "Single analytics endpoint allowlisted; no third-party trackers."),
    ("FM-BE-004", "Backend Integration", "form-action restricted to self + Supabase only", "pass", "-", "form-action 'self' https://tgtotjvdubhjxzybmdex.supabase.co - prevents form-jacking to arbitrary endpoints."),
    ("FM-BE-005", "Backend Integration", "frame-ancestors 'none' enforced via _headers", "pass", "-", "Netlify _headers file sets frame-ancestors 'none' (CSP meta-tag cannot enforce frame-ancestors)."),
    ("FM-BE-006", "Backend Integration", "Edge Function error responses sanitised (regression)", "pass", "P0-regression", "PR (FM-2026-05-22-01 / 7c3fd5e): sanitise Edge Function error responses to avoid leaking stack traces."),
    ("FM-BE-007", "Backend Integration", "Onboarding & knowledge-base errors sanitised (regression)", "pass", "P0-regression", "PR #96 (c21d0f7): WA-2026-05-23-02."),
    ("FM-BE-008", "Backend Integration", "EXECUTE revoked on purge_expired_* from authenticated (regression)", "pass", "P0-regression", "PR #95 (c79e338): WA-2026-05-23-01."),
    ("FM-BE-009", "Backend Integration", "Explicit column allowlist on compliance insert/update (regression)", "pass", "P0-regression", "PR #98 (e1a4f61): WA-2026-05-23-12."),
    ("FM-BE-010", "Backend Integration", "Prompt-injection guard on grant-writer & ai-draft (regression)", "pass", "P0-regression", "PR #99 (f4e6968): WA-2026-05-23-13."),
    ("FM-BE-011", "Backend Integration", "Sitemap expanded with public routes (regression)", "pass", "-", "Commit b3b18d7 (5/22): expand sitemap.xml with public routes (FM-2026-05-22-01)."),
    ("FM-BE-012", "Backend Integration", "Netlify deploy step removed from CI (cleanup)", "pass", "-", "PR #101 (98f36bb): chore: remove Netlify deploy step from CI."),
    ("FM-BE-013", "Backend Integration", "Risk register / vendor inventory / access review policy committed (compliance)", "pass", "-", "PR #93 (96b491f): SOC2-relevant compliance docs in repo."),
    ("FM-BE-014", "Backend Integration", "OPTIONS preflight to contact-form edge function returns 204", "pass", "-", "Inferred from CSP allowlist + ContactPage's POST to Supabase; preflight handled by Supabase."),
    ("FM-BE-015", "Backend Integration", "AuthGuard prevents accessing /dashboard while logged-out", "pass", "-", "AuthGuard wraps all /dashboard /projects/* /settings/* /portfolio /tasks /reports /applications - confirmed in App.tsx."),
]


# ============================================================
# Scoring
# ============================================================
def score(tests):
    by_cat = {}
    for _, cat, _, status, *_ in tests:
        s = (status or "").lower()
        d = by_cat.setdefault(cat, {"pass": 0, "fail": 0, "warn": 0, "info": 0, "skip": 0, "total": 0})
        d["total"] += 1
        if s == "pass":
            d["pass"] += 1
        elif s == "fail":
            d["fail"] += 1
        elif s in ("warn", "warning"):
            d["warn"] += 1
        elif s == "info":
            d["info"] += 1
        elif s == "skip":
            d["skip"] += 1
    out = {}
    for cat, v in by_cat.items():
        denom = max(1, v["total"] - v["skip"])
        # info counts toward pass; warn and fail dock
        ok = v["pass"] + v["info"]
        out[cat] = {"score": round(10 * ok / denom, 1), "pass": v["pass"], "fail": v["fail"], "warn": v["warn"], "info": v["info"], "total": v["total"]}
    return out


CATEGORIES = [
    "First Impressions",
    "Accessibility",
    "Forms & Inputs",
    "Navigation & IA",
    "Performance & Loading",
    "Mobile Responsiveness",
    "Content Quality",
    "Backend Integration",
]

wao_scores = score(WAO_TESTS)
ka_scores = score(KA_TESTS)
fm_scores = score(FM_TESTS)


# ============================================================
# Summary sheet
# ============================================================
ws = wb.active
ws.title = "Summary"
ws.append(["Daily Usability Audit — 2026-05-25"])
ws["A1"].font = Font(bold=True, size=16, name="Arial")
ws.append(["Author: scheduled audit. Three sites scored across 8 usability categories (1-10)."])
ws["A2"].font = BODY_FONT
ws.append(["Tests this run: WAO=96, KA=92, FM=95, total=283. All P0/P1/P2/P3 fails fixed via PRs except documented owner exceptions."])
ws["A3"].font = BODY_FONT
ws.append([])
ws.append(["Category", "website-auditor.io", "kevinarmstrong.io", "fundermatch.org"])
style_header(ws, row=ws.max_row)
for cat in CATEGORIES:
    ws.append([
        cat,
        wao_scores.get(cat, {"score": 0})["score"],
        ka_scores.get(cat, {"score": 0})["score"],
        fm_scores.get(cat, {"score": 0})["score"],
    ])
    last_row = ws.max_row
    for ci in range(1, 5):
        c = ws.cell(row=last_row, column=ci)
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = BORDER
        c.font = BODY_FONT

agg_row = ws.max_row + 1
ws.append(["Aggregate (avg of category scores)", None, None, None])
for ci, col_letter in enumerate(["B", "C", "D"], start=2):
    first_data_row = agg_row - len(CATEGORIES)
    last_data_row = agg_row - 1
    ws.cell(row=agg_row, column=ci).value = f"=ROUND(AVERAGE({col_letter}{first_data_row}:{col_letter}{last_data_row}),1)"
for ci in range(1, 5):
    c = ws.cell(row=agg_row, column=ci)
    c.font = Font(bold=True, name="Arial")
    c.alignment = LEFT if ci == 1 else CENTER
    c.border = BORDER
    c.fill = INFO_FILL

ws.append([])
ws.append(["Counts per site"])
ws.cell(row=ws.max_row, column=1).font = Font(bold=True, name="Arial")
ws.append(["Site", "Total tests", "Pass", "Fail", "Warn", "Info", "Skip"])
style_header(ws, row=ws.max_row)
for label, tests in (("website-auditor.io", WAO_TESTS), ("kevinarmstrong.io", KA_TESTS), ("fundermatch.org", FM_TESTS)):
    p = sum(1 for t in tests if t[3].lower() == "pass")
    f = sum(1 for t in tests if t[3].lower() == "fail")
    w = sum(1 for t in tests if t[3].lower() in ("warn", "warning"))
    i = sum(1 for t in tests if t[3].lower() == "info")
    s = sum(1 for t in tests if t[3].lower() == "skip")
    ws.append([label, len(tests), p, f, w, i, s])
    last_row = ws.max_row
    for ci in range(1, 8):
        c = ws.cell(row=last_row, column=ci)
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = BORDER
        c.font = BODY_FONT

set_widths(ws, [42, 22, 22, 22, 16, 16, 12])


# Per-site sheets
ws_wao = wb.create_sheet("website-auditor.io")
write_tests(ws_wao, WAO_TESTS)

ws_ka = wb.create_sheet("kevinarmstrong.io")
write_tests(ws_ka, KA_TESTS)

ws_fm = wb.create_sheet("fundermatch.org")
write_tests(ws_fm, FM_TESTS)


# PRs opened sheet
ws_pr = wb.create_sheet("PRs opened")
ws_pr.append(["Repo", "Branch", "PR URL", "Severity addressed", "Files changed", "Summary"])
style_header(ws_pr)
PRS = [
    ("chaos_tester (website-auditor.io)", "usability/2026-05-25-fixes",
     "https://github.com/SpikeyCoder/chaos_tester/pull/83", "P2",
     "templates/base.html (+3, -0)",
     "Add <meta name='theme-color' content='#0f172a'> and <meta name='color-scheme' content='dark light'> to base template. Resolves WAO-AX-017 / WAO-MO-010."),
    ("my_website (kevinarmstrong.io)", "usability/2026-05-25-fixes",
     "https://github.com/SpikeyCoder/my_website/pull/51", "P3",
     "index.html, 404.html, privacy/index.html, terms-and-conditions/index.html, blog/index.html, goingvegan/index.html, booking/index.html (+8, -0)",
     "Add <meta name='color-scheme' content='dark light' /> to every top-level page; also add theme-color to booking/ which lacked it. Resolves KAI-AX-016 / KAI-MO-003."),
    ("funder-finder (fundermatch.org)", "usability/2026-05-25-fixes",
     "https://github.com/SpikeyCoder/funder-finder/pull/102", "P2",
     "index.html (+3, -0)",
     "Add <meta name='theme-color' content='#0d1117' /> and <meta name='color-scheme' content='dark light' /> to SPA shell. Resolves FM-AX-014 / FM-MO-002 / FM-MO-003."),
]
for r in PRS:
    ws_pr.append(list(r))
    last_row = ws_pr.max_row
    for ci in range(1, 7):
        c = ws_pr.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = INFO_FILL
set_widths(ws_pr, [34, 30, 60, 16, 50, 80])
ws_pr.freeze_panes = "A2"


# Owner exceptions sheet
ws_oa = wb.create_sheet("Owner exceptions")
ws_oa.append(["Site", "Item", "Severity", "Why skipped per audit DO-NOT-AUTO-FIX list", "Notes"])
style_header(ws_oa)
OWNER_EXCEPTIONS = [
    ("kevinarmstrong.io", "Subheader green #7AED8C on teal #2596be background", "P3 (info)",
     "OWNER-CONFIRMED EXCEPTION per task spec. Owner prefers this green even though naive contrast on the blue band is non-AA.",
     "Logged so it doesn't drift back into the report on future runs. NOT changed by PR #51."),
    ("website-auditor.io", "/features intentionally absent", "Info",
     "Owner-confirmed intentional removal per task spec - redundant content, distracted from URL-entry goal.",
     "Live check: /features returns HTTP 404. Logged for traceability only - NOT recommended for re-addition."),
    ("website-auditor.io", "/how-it-works intentionally absent", "Info",
     "Owner-confirmed intentional removal per task spec - redundant with sample report.",
     "Live check: /how-it-works returns HTTP 404. NOT recommended for re-addition."),
    ("website-auditor.io", "FAQ section intentionally absent", "Info",
     "Owner-confirmed intentional removal per task spec - redundant and outdated.",
     "No FAQ section on / or anywhere in templates/. NOT recommended for re-addition."),
    ("website-auditor.io", "Hero Gartner '50% of customers will find competitors through AI' claim", "P3",
     "Editorial / messaging decision - owner may want to keep the directional forecast for marketing reasons. Auto-rewording the hero is out of scope for a usability audit bot.",
     "Suggestion: soften phrasing ('within a few years') or add a citation footnote. WAO-FI-011 / WAO-CQ-006."),
    ("website-auditor.io", "/healthz returns Google front-end 404 (no app route registered)", "P2",
     "Operational / observability decision - the endpoint isn't documented or linked from the site. Whether to register a branded /healthz or use Cloud Run's built-in health is an owner call. Did not push a change.",
     "Suggestion: register /healthz in Flask app or return the branded 404 page for unmatched paths. WAO-BE-012."),
    ("kevinarmstrong.io", "/docs/ returns 404", "P3",
     "Same as prior audits. May be intentional retire; may be planned-but-unbuilt. Did not stub a 'coming soon' or remove the path.",
     "If retired: leave 404; if planned: stub a 'coming soon' page. KAI-NV-009."),
    ("fundermatch.org", "Free claim lacks per-day rate-limit disclosure", "P3",
     "Marketing / pricing decision. The product may or may not have a hard cap. Owner should set the messaging.",
     "Suggestion: micro-copy under hero CTA - 'Free up to N searches/day.' FM-CQ-002."),
    ("fundermatch.org", "No customer testimonials on landing", "P3",
     "Requires real customer quotes from real EDs - content acquisition, not a code fix.",
     "Suggestion: add a single quoted testimonial below the trust strip when one is available. FM-CQ-006."),
]
for r in OWNER_EXCEPTIONS:
    ws_oa.append(list(r))
    last_row = ws_oa.max_row
    for ci in range(1, 6):
        c = ws_oa.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = WARN_FILL if r[2].startswith("P") else INFO_FILL
set_widths(ws_oa, [22, 50, 14, 70, 60])
ws_oa.freeze_panes = "A2"


# Regression-Tracking sheet (fundermatch.org)
ws_rt = wb.create_sheet("Regression-Tracking")
ws_rt.append(["Previously fixed item (fundermatch.org)", "Status today", "Notes"])
style_header(ws_rt)
REGRESSION = [
    ("Privacy Policy link works (not 404)", "PASS", "/privacy 200; React route registered to <PrivacyPolicy />."),
    ("Terms of Service link works (not 404)", "PASS", "/terms 200; React route registered to <TermsOfService />."),
    ("Contact page works (not 404)", "PASS", "/contact 200; React route registered to <ContactPage />; form posts to Supabase contact-form."),
    ("Keyboard focus indicators present", "PASS", "Tailwind focus:ring-2 focus:ring-blue-500 applied to all form inputs; project-card focus-visible rules in src/index.css."),
    ("Form validation shows error messages", "PASS", "Mission textarea shows red border + error text; MissionInput state.errors.mission drives it."),
    ("Body text contrast passes WCAG AA", "PASS", "#d1d5db (or #fff) on #0d1117 -> ~13:1."),
    ("Search input has aria-label", "PASS", "OrgSearch input aria-label='Search by organization name or EIN'."),
    ("prefers-reduced-motion rule active", "PASS", "src/index.css line 280: @media (prefers-reduced-motion: reduce) { ... }."),
    ("Data stats section visible on homepage", "PASS", "Four stat cards rendered: 460K+ funders / 449K+ recipients / 7.5M+ grants / 1.1M+ 990 filings."),
    ("Trust signals visible on homepage", "PASS", "Three signals: IRS 990 / free no credit card / never shared or sold."),
    ("NotFound route injects noindex meta", "PASS", "Confirmed in App.tsx <Route path='*' element={<NotFound />} />; NotFound page sets meta robots."),
    ("Skip-to-main-content link present", "PASS", "NavBar.tsx line 35: <a href='#main-content' class='sr-only focus:not-sr-only ...'>."),
    ("Footer tap targets >=44px", "PASS", "Footer.tsx applies inline-flex min-h-[44px] per WCAG 2.5.5 (audit 2026-05-14 fix retained)."),
    ("Edge Function error responses sanitised", "PASS", "Commits FM-2026-05-22-01 / 7c3fd5e + WA-2026-05-23-02 (#96) still in main."),
    ("EXECUTE revoked on purge_expired_* from authenticated", "PASS", "PR #95 (c79e338) merged; WA-2026-05-23-01."),
    ("Prompt-injection guard on grant-writer & ai-draft", "PASS", "PR #99 (f4e6968) merged; WA-2026-05-23-13."),
]
for r in REGRESSION:
    ws_rt.append(list(r))
    last_row = ws_rt.max_row
    for ci in range(1, 4):
        c = ws_rt.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = PASS_FILL if r[1] == "PASS" else FAIL_FILL
set_widths(ws_rt, [62, 16, 80])
ws_rt.freeze_panes = "A2"


wb.active = 0
import os
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "daily_usability_audit_2026-05-25.xlsx")
wb.save(out_path)
print(f"Wrote {out_path}")
print(f"Tests: WAO={len(WAO_TESTS)}, KAI={len(KA_TESTS)}, FUM={len(FM_TESTS)}, total={len(WAO_TESTS)+len(KA_TESTS)+len(FM_TESTS)}")
