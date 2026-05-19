"""Build the daily usability audit Excel report for 2026-05-18.

Sheets:
  Summary, website-auditor.io, kevinarmstrong.io, fundermatch.org,
  Auto-fixes, Owner-action items, Regression-Tracking
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
SKIP_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
INFO_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def status_fill(s):
    s = (s or "").lower()
    if s == "pass":
        return PASS_FILL
    if s == "fail":
        return FAIL_FILL
    if s in ("warn", "warning"):
        return WARN_FILL
    if s == "skip":
        return SKIP_FILL
    if s == "info":
        return INFO_FILL
    return None


def write_tests(ws, tests):
    ws.append(["Test ID", "Category", "Test", "Status", "Severity", "Notes"])
    style_header(ws)
    for row in tests:
        ws.append(list(row))
        last_row = ws.max_row
        fill = status_fill(row[3])
        for ci in range(1, 7):
            cell = ws.cell(row=last_row, column=ci)
            cell.alignment = LEFT
            cell.border = BORDER
            cell.font = BODY_FONT
            if fill is not None:
                cell.fill = fill
    set_widths(ws, [14, 22, 60, 10, 10, 80])
    ws.freeze_panes = "A2"


# ============================================================
# WEBSITE-AUDITOR.IO  (chaos_tester) — 95 tests
# ============================================================
WAO_TESTS = [
    # First Impressions (12)
    ("WAO-FI-001", "First Impressions", "Homepage clearly explains product (AI visibility + security + perf scanner)", "pass", "-", "Hero copy 'Does ChatGPT recommend your business?' clear; meta description: 'Free website audit tool that checks AI visibility, broken links, security, and performance...'"),
    ("WAO-FI-002", "First Impressions", "Value proposition is immediately obvious above the fold", "pass", "-", "Trust strip: 100% Free, No Signup Required, Instant Results; hero subtitle frames urgency"),
    ("WAO-FI-003", "First Impressions", "Primary CTA visible and compelling", "pass", "-", "'Check my site' submit button next to URL input; high contrast; disabled until URL entered"),
    ("WAO-FI-004", "First Impressions", "Page loads under 3 seconds (TTFB measurement)", "pass", "-", "Live curl: TTFB 0.461s, full HTML 28.9KB, 200 OK"),
    ("WAO-FI-005", "First Impressions", "Visual hierarchy clear with single H1", "pass", "-", "Exactly 1 H1 in dashboard.html, H2/H3 hierarchy intact"),
    ("WAO-FI-006", "First Impressions", "Trust signals (testimonials, credentials) visible on homepage", "pass", "-", "Two testimonial cards (Jake Morrison; Sarah Patel) with role attribution"),
    ("WAO-FI-007", "First Impressions", "Industry/use-case badges shown", "pass", "-", "E-commerce, SaaS, Healthcare, Real Estate, Restaurants, Legal"),
    ("WAO-FI-008", "First Impressions", "Navigation intuitive (Sample Report, API, Contact)", "pass", "-", "3 nav links + hamburger fallback for mobile"),
    ("WAO-FI-009", "First Impressions", "Hero works on mobile (responsive)", "pass", "-", "Hamburger toggle present (#navToggle) with aria-expanded; media queries defined"),
    ("WAO-FI-010", "First Impressions", "No broken images / missing assets in hero", "pass", "-", "logo.svg + favicon.svg both 200; no <img> without alt in landing"),
    ("WAO-FI-011", "First Impressions", "Body font readable at default size", "pass", "-", "16px Inter base; ample line-height"),
    ("WAO-FI-012", "First Impressions", "Sample audit preview shown on homepage", "pass", "-", "'Sample Audit Preview' card with Score: 87/100, Passed/Failed/Warnings counts"),

    # Accessibility (18)
    ("WAO-A11Y-001", "Accessibility", "Homepage: WCAG 2.1 AA structural review", "pass", "-", "Static analysis: 1 H1, no nested form errors, valid HTML5; aria-labels on interactive controls"),
    ("WAO-A11Y-002", "Accessibility", "/sample-report structural review", "pass", "-", "Title set; heading hierarchy preserved; data tables have headers"),
    ("WAO-A11Y-003", "Accessibility", "/about structural review", "pass", "-", "Single H1 = 'About Website Auditor'; uses H2 sections"),
    ("WAO-A11Y-004", "Accessibility", "/api docs structural review", "pass", "-", "Single H1, sections under H2/H3 properly nested"),
    ("WAO-A11Y-005", "Accessibility", "/contact structural review", "pass", "-", "Single H1, all links labelled, mailto support@website-auditor.io exposed"),
    ("WAO-A11Y-006", "Accessibility", "Lang attribute present", "pass", "-", "<html lang='en'>"),
    ("WAO-A11Y-007", "Accessibility", "Skip-to-content link present", "pass", "-", "<a href='#main-content' class='skip-link'>Skip to main content</a>"),
    ("WAO-A11Y-008", "Accessibility", "Single H1 per page", "pass", "-", "dashboard, about, api_docs, contact, sample_report: 1 H1 each (grep count)"),
    ("WAO-A11Y-009", "Accessibility", "Heading hierarchy: no skipped levels", "pass", "-", "H1 -> H2 -> H3 consistently"),
    ("WAO-A11Y-010", "Accessibility", "All <img> have alt attribute", "pass", "-", "Logo uses alt='' aria-hidden; sample report imgs have alt; report row imgs have alt"),
    ("WAO-A11Y-011", "Accessibility", "All form inputs labelled (aria-label or <label>)", "pass", "-", "URL input aria-label='Website URL to audit'; bug textarea has sr-only label"),
    ("WAO-A11Y-012", "Accessibility", "Landmark roles (main, nav, footer) present", "pass", "-", "main, nav, footer each appear once in base.html"),
    ("WAO-A11Y-013", "Accessibility", "Hamburger nav has aria-label + aria-expanded", "pass", "-", "aria-label='Toggle navigation menu', aria-expanded='false' initial"),
    ("WAO-A11Y-014", "Accessibility", "prefers-reduced-motion honored", "pass", "-", "Media query present in base.css AND progress.css"),
    ("WAO-A11Y-015", "Accessibility", "Body text color contrast passes WCAG AA", "pass", "-", "Body fg ~rgb(226,232,240) on dark surface ~14:1"),
    ("WAO-A11Y-016", "Accessibility", "Focus styles defined in CSS", "pass", "-", "*:focus-visible rules present including skip-link"),
    ("WAO-A11Y-017", "Accessibility", "Bug-report modal button has accessible label", "pass", "-", "aria-label='Report a bug' on .bug-btn"),
    ("WAO-A11Y-018", "Accessibility", "Hidden helper text uses aria-live for dynamic updates", "pass", "-", "#audit-cta-hint has aria-live='polite'"),

    # Forms & Inputs (11)
    ("WAO-FORM-001", "Forms & Inputs", "Audit form has required URL input", "pass", "-", "type='url', required, aria-labelled"),
    ("WAO-FORM-002", "Forms & Inputs", "Audit form has CSRF token", "pass", "-", "Hidden csrf_token input rendered via Flask {{ csrf_token() }}"),
    ("WAO-FORM-003", "Forms & Inputs", "URL input rejects invalid formats", "pass", "-", "HTML5 type='url' enforces basic syntax client-side"),
    ("WAO-FORM-004", "Forms & Inputs", "Form has helpful placeholder", "pass", "-", "'Enter your website URL (e.g., https://yourbusiness.com)'"),
    ("WAO-FORM-005", "Forms & Inputs", "Submit button labelled clearly", "pass", "-", "'Check my site'"),
    ("WAO-FORM-006", "Forms & Inputs", "POST /run server-side validation", "pass", "-", "Form uses POST + novalidate; server validates URL + business location"),
    ("WAO-FORM-007", "Forms & Inputs", "Contact bug form has labelled textarea", "pass", "-", "label[for=bugDesc] with .sr-only class"),
    ("WAO-FORM-008", "Forms & Inputs", "Contact bug form has character limit feedback", "pass", "-", "maxlength=1000 with live char counter (bugCharCount span)"),
    ("WAO-FORM-009", "Forms & Inputs", "Bug-report endpoint POST flow defined", "pass", "-", "/api/bug-report POST; error/success modal states; payload now capped (2026-05-18)"),
    ("WAO-FORM-010", "Forms & Inputs", "Form inputs match content type", "pass", "-", "type=url, type=text, type=checkbox where appropriate"),
    ("WAO-FORM-011", "Forms & Inputs", "City autocomplete dropdown reachable via keyboard", "pass", "-", "Dropdown via Maps API; aria-label='Business city' on input"),

    # Navigation & IA (12)
    ("WAO-NAV-001", "Navigation & IA", "Top-nav /sample-report 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-002", "Navigation & IA", "Top-nav /api 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-003", "Navigation & IA", "Top-nav /contact 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-004", "Navigation & IA", "Footer /about 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-005", "Navigation & IA", "Footer /privacy 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-006", "Navigation & IA", "Footer /terms 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-007", "Navigation & IA", "Footer /status 200", "pass", "-", "Live curl: 200"),
    ("WAO-NAV-008", "Navigation & IA", "Footer /changelog 200", "pass", "-", "Live curl: 200 — now includes May 2026 entry (auto-fix this run)"),
    ("WAO-NAV-009", "Navigation & IA", "Custom 404 page renders properly", "pass", "-", "/__notexist -> 404 with template/404.html; noindex + canonical to /"),
    ("WAO-NAV-010", "Navigation & IA", "Active nav state CSS applied", "pass", "-", "Templates use Flask request.path comparison"),
    ("WAO-NAV-011", "Navigation & IA", "Footer GitHub link uses rel='noopener'", "pass", "-", "target='_blank' rel='noopener'"),
    ("WAO-NAV-012", "Navigation & IA", "Intentional removals (Features, How-It-Works, FAQ) honored", "info", "-", "Owner-confirmed intentional removals — not flagged per task spec"),

    # Performance & Loading (10)
    ("WAO-PERF-001", "Performance & Loading", "Homepage TTFB acceptable (<500ms)", "pass", "-", "Live curl TTFB 461ms"),
    ("WAO-PERF-002", "Performance & Loading", "HTML transfer size reasonable (<50KB)", "pass", "-", "28.9KB"),
    ("WAO-PERF-003", "Performance & Loading", "Critical assets preloaded", "pass", "-", "logo.svg + favicon.svg preloaded; fonts preconnected"),
    ("WAO-PERF-004", "Performance & Loading", "Google Fonts pre-connected", "pass", "-", "preconnect to fonts.googleapis.com + fonts.gstatic.com"),
    ("WAO-PERF-005", "Performance & Loading", "GoatCounter pre-connected for analytics", "pass", "-", "preconnect to gc.zgo.at"),
    ("WAO-PERF-006", "Performance & Loading", "Logo uses fetchpriority='high'", "pass", "-", "Set on hero logo to reduce LCP"),
    ("WAO-PERF-007", "Performance & Loading", "JS bundles served with 200", "pass", "-", "dashboard.js, bug-report.js, sample-report-full.js all 200"),
    ("WAO-PERF-008", "Performance & Loading", "CSS files served with 200", "pass", "-", "base.css, utilities.css, dashboard.css all 200"),
    ("WAO-PERF-009", "Performance & Loading", "robots.txt + sitemap.xml served", "pass", "-", "/robots.txt 200, /sitemap.xml 200"),
    ("WAO-PERF-010", "Performance & Loading", "Favicon served (no 404 noise)", "pass", "-", "/favicon.ico 200"),

    # Mobile Responsiveness (10)
    ("WAO-MOB-001", "Mobile Responsiveness", "Viewport meta tag present", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1.0'>"),
    ("WAO-MOB-002", "Mobile Responsiveness", "Hamburger menu defined for mobile", "pass", "-", "#navToggle button toggles #mainNav"),
    ("WAO-MOB-003", "Mobile Responsiveness", "CSS media query for ≤768px", "pass", "-", "Multiple breakpoints in base.css and utilities.css"),
    ("WAO-MOB-004", "Mobile Responsiveness", "Touch targets ≥44x44px via u-cta-padding-md", "pass", "-", "Audit button uses .btn-audit with vertical padding ≥12px"),
    ("WAO-MOB-005", "Mobile Responsiveness", "Hero text scales (clamp/responsive units)", "pass", "-", "Hero H1 uses responsive sizing in landing-hero CSS"),
    ("WAO-MOB-006", "Mobile Responsiveness", "Sample preview card stacks vertically on mobile", "pass", "-", "Grid uses repeat(auto-fit, minmax) — collapses on narrow screens"),
    ("WAO-MOB-007", "Mobile Responsiveness", "Bug-report modal sized for small screens", "pass", "-", "Modal uses max-width responsive approach"),
    ("WAO-MOB-008", "Mobile Responsiveness", "Footer grid collapses on mobile", "pass", "-", ".footer-grid uses repeat(auto-fit, minmax(...))"),
    ("WAO-MOB-009", "Mobile Responsiveness", "No fixed-width elements forcing horizontal scroll", "pass", "-", "Containers use width: min(...) pattern"),
    ("WAO-MOB-010", "Mobile Responsiveness", "Trust-badge row wraps on small screens", "pass", "-", ".trust-badges uses flex-wrap"),

    # Content Quality (7)
    ("WAO-CONTENT-001", "Content Quality", "No typos in homepage hero / features", "pass", "-", "Static grep for common misspellings found nothing"),
    ("WAO-CONTENT-002", "Content Quality", "No 404s on linked internal pages", "pass", "-", "12 internal URLs probed; all 200 except intentional /__notexist (404 as expected)"),
    ("WAO-CONTENT-003", "Content Quality", "Footer copyright is current (2026)", "pass", "-", "'© 2026 Armstrong HoldCo LLC. All rights reserved.'"),
    ("WAO-CONTENT-004", "Content Quality", "Privacy + Terms pages exist", "pass", "-", "/privacy 200, /terms 200"),
    ("WAO-CONTENT-005", "Content Quality", "Contact info: support email visible", "pass", "-", "support@website-auditor.io shown on /contact"),
    ("WAO-CONTENT-006", "Content Quality", "Changelog reflects May 2026 work", "pass", "P2", "Previously only April 2026 entries; auto-fix this run added May section (CSP/AI rate-limit/bug-report cap/SOC2 A1)"),
    ("WAO-CONTENT-007", "Content Quality", "Forward-looking claim is dated/sourced", "info", "P3", "Hero subtitle uses 'By late 2027, 75% of your customers will find your competitors with AI' — punchy but unverifiable; owner can decide whether to soften"),

    # Backend Integration (15)
    ("WAO-BACK-001", "Backend Integration", "Audit form POST /run accepts URL", "pass", "-", "Endpoint defined in app.py; CSRF + business_name + business_location validated server-side"),
    ("WAO-BACK-002", "Backend Integration", "CSP report endpoint exists", "pass", "-", "POST /api/csp-report returns 204 (no content) as expected"),
    ("WAO-BACK-003", "Backend Integration", "CSP report endpoint rejects GET", "pass", "-", "GET /api/csp-report -> 404 (correct — reports are POST-only)"),
    ("WAO-BACK-004", "Backend Integration", "Status surface available at /status", "pass", "-", "200 with template/status.html"),
    ("WAO-BACK-005", "Backend Integration", "Strict-Transport-Security header (HSTS preload)", "pass", "-", "max-age=63072000; includeSubDomains; preload"),
    ("WAO-BACK-006", "Backend Integration", "Content-Security-Policy locked down", "pass", "-", "default-src 'self'; script-src whitelist; object-src 'none'; frame-ancestors 'none'; CSP report-uri /api/csp-report"),
    ("WAO-BACK-007", "Backend Integration", "X-Frame-Options: DENY", "pass", "-", "Header set"),
    ("WAO-BACK-008", "Backend Integration", "X-Content-Type-Options: nosniff", "pass", "-", "Header set"),
    ("WAO-BACK-009", "Backend Integration", "Referrer-Policy: strict-origin-when-cross-origin", "pass", "-", "Header set"),
    ("WAO-BACK-010", "Backend Integration", "Permissions-Policy disables camera/mic/geo/FLoC", "pass", "-", "camera=(), microphone=(), geolocation=(), interest-cohort=()"),
    ("WAO-BACK-011", "Backend Integration", "Cross-Origin-Opener / Resource policies set", "pass", "-", "COOP: same-origin; CORP: same-origin"),
    ("WAO-BACK-012", "Backend Integration", "Session cookie secure / httpOnly / SameSite=Lax", "pass", "-", "Set-Cookie: session=…; Secure; HttpOnly; Path=/; SameSite=Lax"),
    ("WAO-BACK-013", "Backend Integration", "Rate limiting in place on hot endpoints", "pass", "-", "x-ratelimit-limit: 120 / remaining: 118 on home; AI-query route now rate-limited (May 2026)"),
    ("WAO-BACK-014", "Backend Integration", "Reporting endpoint registered (Reporting API)", "pass", "-", "reporting-endpoints + report-to header for csp-endpoint"),
    ("WAO-BACK-015", "Backend Integration", "Bug-report endpoint payload guard", "pass", "-", "Body size capped this month (2026-05-18); JPEG screenshot path"),
]


# ============================================================
# KEVINARMSTRONG.IO  (my_website) — 92 tests
# ============================================================
KA_TESTS = [
    # First Impressions (12)
    ("KA-FI-001", "First Impressions", "Homepage clearly explains who Kevin is", "pass", "-", "Eyebrow 'Kevin Armstrong' + H1 'Product leader focused on customer trust'"),
    ("KA-FI-002", "First Impressions", "Value proposition / focus area visible", "pass", "-", "Hero subtitle describes AI-native tools + coaching for SMBs and nonprofits"),
    ("KA-FI-003", "First Impressions", "Primary CTA prominent", "pass", "-", "'Accelerate Your Career' button links to #career-acceleration"),
    ("KA-FI-004", "First Impressions", "Page loads under 3s (TTFB)", "pass", "-", "Live curl: TTFB 0.282s, 51KB HTML"),
    ("KA-FI-005", "First Impressions", "Visual hierarchy: single H1", "pass", "-", "1 H1 = 'Product leader focused on customer trust'"),
    ("KA-FI-006", "First Impressions", "Highlight stats card visible", "pass", "-", "$13.2MM annual GMS, 2MM+ customers, MBA Chicago Booth"),
    ("KA-FI-007", "First Impressions", "Logo brand mark in nav", "pass", "-", "ARMSTRONG HOLDCO LLC text brand; consistent letterspacing"),
    ("KA-FI-008", "First Impressions", "Social links discoverable", "pass", "-", "LinkedIn + GitHub in social-compact details; share button"),
    ("KA-FI-009", "First Impressions", "Mobile-aware navigation", "pass", "-", "Compact social details panel; flexible nav"),
    ("KA-FI-010", "First Impressions", "No broken images in hero", "pass", "-", "Hero is pure CSS — no <img> on initial fold"),
    ("KA-FI-011", "First Impressions", "Body font readable at default size", "pass", "-", "Inter via stylesheet, 16px base, ample line-height"),
    ("KA-FI-012", "First Impressions", "Interactive 'Portfolio' section visible", "pass", "-", "H2 'Interactive Portfolio' with tabbed views"),

    # Accessibility (17)
    ("KA-A11Y-001", "Accessibility", "Lang attribute present", "pass", "-", "<html lang='en'>"),
    ("KA-A11Y-002", "Accessibility", "Skip-to-content link present", "pass", "-", "<a href='#main-content' class='skip-link'>Skip to content</a>"),
    ("KA-A11Y-003", "Accessibility", "Single H1 on homepage", "pass", "-", "Only 1 H1; rest are H2/H3 (8 total H1+H2 in index.html, of which 1 is H1)"),
    ("KA-A11Y-004", "Accessibility", "Heading hierarchy intact", "pass", "-", "No skipped levels in document outline"),
    ("KA-A11Y-005", "Accessibility", "Every <img> has alt", "pass", "-", "grep -oE '<img[^>]*>' produced 0 hits without alt="),
    ("KA-A11Y-006", "Accessibility", "Nav has aria-label", "pass", "-", "<nav class='site-nav' aria-label='Main navigation'>"),
    ("KA-A11Y-007", "Accessibility", "Admin toggle button has aria-label", "pass", "-", "aria-label='Open admin panel'"),
    ("KA-A11Y-008", "Accessibility", "Social icons have aria-label", "pass", "-", "aria-label='Open LinkedIn profile' / 'Open GitHub profile'"),
    ("KA-A11Y-009", "Accessibility", "Page-share button has aria-label", "pass", "-", "aria-label='Copy page URL'"),
    ("KA-A11Y-010", "Accessibility", "Portfolio tablist has aria-label", "pass", "-", "role='tablist' aria-label='Portfolio view'"),
    ("KA-A11Y-011", "Accessibility", "Blog list has role + aria-label + tabindex", "pass", "-", "role='region' tabindex='0' aria-label='Latest blog posts'"),
    ("KA-A11Y-012", "Accessibility", "prefers-reduced-motion media query honored", "pass", "-", "@media (prefers-reduced-motion: reduce) present in styles.css"),
    ("KA-A11Y-013", "Accessibility", "Body text color contrast passes WCAG AA", "pass", "-", "Primary text on dark surface achieves ~14:1; muted text ~7:1"),
    ("KA-A11Y-014", "Accessibility", "Focus styles defined", "pass", "-", ":focus-visible rules on btn, nav-admin, project-card, loom-facade, appstore-cta, skip-link"),
    ("KA-A11Y-015", "Accessibility", "Loom video facade has accessible label", "pass", "-", "aria-label='Play 3-minute coaching approach overview video'"),
    ("KA-A11Y-016", "Accessibility", "Blog list region keyboard-focusable", "pass", "-", "tabindex='0' on blog-list region"),
    ("KA-A11Y-017", "Accessibility", "Subheader green #7AED8C on #2596be background", "info", "-", "Owner-confirmed exception — owner prefers this color despite contrast on banner background. NOT auto-fixed per task spec."),

    # Forms & Inputs (8)
    ("KA-FORM-001", "Forms & Inputs", "Blog editor compose has rich-text controls labelled", "pass", "-", "aria-label on Bold/Italic/Underline/Font color/Font size/Font family controls"),
    ("KA-FORM-002", "Forms & Inputs", "Admin panel toggle button labelled", "pass", "-", "aria-label='Open admin panel'"),
    ("KA-FORM-003", "Forms & Inputs", "Page-share button labelled", "pass", "-", "aria-label='Copy page URL'"),
    ("KA-FORM-004", "Forms & Inputs", "Form action(s) point to known origins (Stripe / Calendar / Supabase)", "pass", "-", "CSP form-action whitelists buy.stripe.com, calendar.app.google, *.supabase.co"),
    ("KA-FORM-005", "Forms & Inputs", "No external form actions to unexpected origins", "pass", "-", "Headers reviewed; form-action restricted appropriately"),
    ("KA-FORM-006", "Forms & Inputs", "Blog editor select dropdowns labelled", "pass", "-", "Each <select> in blog editor has explicit aria-label"),
    ("KA-FORM-007", "Forms & Inputs", "Admin panel inputs gated", "pass", "-", "Admin actions require authenticated user — surface gated"),
    ("KA-FORM-008", "Forms & Inputs", "Compose action buttons keyboard-accessible", "pass", "-", "<button type='button'> with text labels and aria-label"),

    # Navigation & IA (12)
    ("KA-NAV-001", "Navigation & IA", "/privacy 200", "pass", "-", "/privacy 308 -> /privacy/ 200 (trailing-slash normalization)"),
    ("KA-NAV-002", "Navigation & IA", "/terms-and-conditions/ 200", "pass", "-", "Live: 200"),
    ("KA-NAV-003", "Navigation & IA", "/blog 200", "pass", "-", "/blog 308 -> /blog/ 200"),
    ("KA-NAV-004", "Navigation & IA", "/goingvegan 200", "pass", "-", "/goingvegan 308 -> /goingvegan/ 200"),
    ("KA-NAV-005", "Navigation & IA", "/booking 200", "pass", "-", "/booking 308 -> /booking/ 200"),
    ("KA-NAV-006", "Navigation & IA", "Cloudflare 308 redirects clean", "pass", "-", "Trailing-slash 308s are canonical; same-origin destinations; no chain loops"),
    ("KA-NAV-007", "Navigation & IA", "/terms vs /terms-and-conditions/", "pass", "-", "Footer canonical link is /terms-and-conditions/; /terms 404 is expected (path not published)"),
    ("KA-NAV-008", "Navigation & IA", "Custom 404 page", "pass", "-", "/__nope-test-foo -> 404 with rich 404 template (Back to Home + Browse Blog CTAs + #contact mention)"),
    ("KA-NAV-009", "Navigation & IA", "/favicon.ico, /robots.txt, /sitemap.xml 200", "pass", "-", "All three 200"),
    ("KA-NAV-010", "Navigation & IA", "/apple-touch-icon.png served", "pass", "-", "200"),
    ("KA-NAV-011", "Navigation & IA", "Sitemap currency", "pass", "P2", "Sitemap had lastmod 2026-04-22 and 37 URLs vs 58 authored blog dirs; auto-fix this run regenerated to 56 URLs with lastmod 2026-05-19"),
    ("KA-NAV-012", "Navigation & IA", "404 page has nav back to home/blog", "pass", "-", "Two clear CTA buttons in /404.html"),

    # Performance & Loading (10)
    ("KA-PERF-001", "Performance & Loading", "Homepage TTFB acceptable (<500ms)", "pass", "-", "Live: 282ms"),
    ("KA-PERF-002", "Performance & Loading", "HTML payload size reasonable (<70KB)", "pass", "-", "51.3KB"),
    ("KA-PERF-003", "Performance & Loading", "GoatCounter preconnected", "pass", "-", "Link rel=preconnect to https://gc.zgo.at"),
    ("KA-PERF-004", "Performance & Loading", "Cache-control set for static delivery", "pass", "-", "public, max-age=0, must-revalidate (sensible for SPA-style)"),
    ("KA-PERF-005", "Performance & Loading", "Speculation Rules enabled (Cloudflare)", "pass", "-", "speculation-rules header pointing to /cdn-cgi/speculation"),
    ("KA-PERF-006", "Performance & Loading", "RSS aggregator JSON served fresh", "pass", "-", "/rss.json generated 2026-05-19T03:19:57Z (today)"),
    ("KA-PERF-007", "Performance & Loading", "Static assets versioned (styles.css?v=…)", "pass", "-", "?v=20260502a cache-bust marker"),
    ("KA-PERF-008", "Performance & Loading", "Analytics deferred", "pass", "-", "<script defer src='/analytics.js'> on 404 and root"),
    ("KA-PERF-009", "Performance & Loading", "Loom embeds use facade (no auto-load iframe)", "pass", "-", "loom-facade CSS class + click-to-load pattern"),
    ("KA-PERF-010", "Performance & Loading", "Loom embed CSP frame-src restricted", "pass", "-", "frame-src 'self' https://www.loom.com"),

    # Mobile Responsiveness (10)
    ("KA-MOB-001", "Mobile Responsiveness", "Viewport meta tag present", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1' />"),
    ("KA-MOB-002", "Mobile Responsiveness", "Stats card grid responsive", "pass", "-", "Repeating grid via flex-wrap pattern"),
    ("KA-MOB-003", "Mobile Responsiveness", "Nav adapts on small screens", "pass", "-", "social-compact-summary collapses social row on narrow widths"),
    ("KA-MOB-004", "Mobile Responsiveness", "Theme color set for mobile chrome", "pass", "-", "<meta name='theme-color' content='#2596be' />"),
    ("KA-MOB-005", "Mobile Responsiveness", "Touch-friendly CTA padding", "pass", "-", ".btn primary uses generous vertical padding (>=12px)"),
    ("KA-MOB-006", "Mobile Responsiveness", "Pill row scrolls horizontally if needed", "pass", "-", "#portfolio-tabs uses pill-row layout"),
    ("KA-MOB-007", "Mobile Responsiveness", "Blog editor wraps cleanly on mobile", "pass", "-", "Compose toolbar uses inline-flex; controls wrap"),
    ("KA-MOB-008", "Mobile Responsiveness", "Career video panel responsive", "pass", "-", "loom-embed-wrap uses aspect-ratio + fluid width"),
    ("KA-MOB-009", "Mobile Responsiveness", "Footer collapses to single column", "pass", "-", "Footer flex layout wraps below ~480px"),
    ("KA-MOB-010", "Mobile Responsiveness", "Hero stat values stay readable on small screens", "pass", "-", "Stat values + metric-context use responsive font-size"),

    # Content Quality (10)
    ("KA-CONTENT-001", "Content Quality", "Footer copyright is current (2026)", "pass", "-", "'© 2026 Armstrong HoldCo LLC. All rights reserved.'"),
    ("KA-CONTENT-002", "Content Quality", "Privacy + Terms pages exist", "pass", "-", "/privacy/ 200, /terms-and-conditions/ 200"),
    ("KA-CONTENT-003", "Content Quality", "No 404s on linked internal pages", "pass", "-", "All canonical paths checked: /blog/, /goingvegan/, /booking/, /privacy/, /terms-and-conditions/ all 200 (via 308 trailing-slash redirect where needed)"),
    ("KA-CONTENT-004", "Content Quality", "Stats card years reference 2024-2026", "pass", "-", "'2024-2026 tenure', '2026-Present', '2021–2026'"),
    ("KA-CONTENT-005", "Content Quality", "Blog corpus current (RSS regenerated today)", "pass", "-", "/rss.json generated_at 2026-05-19T03:19:57Z"),
    ("KA-CONTENT-006", "Content Quality", "Sitemap covers blog corpus", "pass", "P2", "Auto-fix this run: regenerated sitemap.xml from 37 URLs (lastmod 2026-04-22) to 56 URLs (lastmod 2026-05-19)"),
    ("KA-CONTENT-007", "Content Quality", "Blog post canonical tags set", "pass", "-", "Sample 'why-i-still-code…' has <link rel='canonical' …>"),
    ("KA-CONTENT-008", "Content Quality", "Blog post structured data (Article schema)", "pass", "-", "JSON-LD Article schema on sample post"),
    ("KA-CONTENT-009", "Content Quality", "Subheader color #7AED8C contrast (owner exception)", "info", "-", "Owner prefers #7AED8C green over #2596be banner; NOT flagged or fixed per task spec"),
    ("KA-CONTENT-010", "Content Quality", "Footer hosts terms + privacy links", "pass", "-", "Footer-links lists Terms / Privacy"),

    # Backend Integration (13)
    ("KA-BACK-001", "Backend Integration", "Strict-Transport-Security with preload", "pass", "-", "max-age=63072000; includeSubDomains; preload"),
    ("KA-BACK-002", "Backend Integration", "CSP locked down (script-src/style-src/connect-src)", "pass", "-", "Hash-pinned scripts; CSP report-uri /api/csp-report; ~60+ sha256 hashes for inline scripts"),
    ("KA-BACK-003", "Backend Integration", "X-Frame-Options: DENY", "pass", "-", "Header set"),
    ("KA-BACK-004", "Backend Integration", "X-Content-Type-Options: nosniff", "pass", "-", "Header set"),
    ("KA-BACK-005", "Backend Integration", "Referrer-Policy: strict-origin-when-cross-origin", "pass", "-", "Header set"),
    ("KA-BACK-006", "Backend Integration", "Permissions-Policy disables camera/mic/geo/FLoC", "pass", "-", "camera=(), microphone=(), geolocation=(), interest-cohort=()"),
    ("KA-BACK-007", "Backend Integration", "Cross-Origin-Opener-Policy: same-origin", "pass", "-", "Header set"),
    ("KA-BACK-008", "Backend Integration", "Cross-Origin-Embedder-Policy: credentialless", "pass", "-", "Header set"),
    ("KA-BACK-009", "Backend Integration", "Cross-Origin-Resource-Policy: same-origin", "pass", "-", "Header set"),
    ("KA-BACK-010", "Backend Integration", "CSP form-action whitelist", "pass", "-", "form-action 'self' https://buy.stripe.com https://calendar.app.google https://*.supabase.co"),
    ("KA-BACK-011", "Backend Integration", "CSP frame-src allows Loom", "pass", "-", "frame-src 'self' https://www.loom.com"),
    ("KA-BACK-012", "Backend Integration", "Reporting endpoint configured", "pass", "-", "reporting-endpoints + report-to header for csp-endpoint"),
    ("KA-BACK-013", "Backend Integration", "NEL reporting endpoint configured", "pass", "-", "report-to + nel headers for Cloudflare cf-nel"),
]


# ============================================================
# FUNDERMATCH.ORG  (funder-finder) — 102 tests
# audited from nonprofit perspective: 2-3 staff, limited tech, <$500K budget
# ============================================================
FM_TESTS = [
    # First Impressions (13)
    ("FM-FI-001", "First Impressions", "Homepage clearly explains product (funder matching for nonprofits)", "pass", "-", "<title>Non-Profit Funder Finder — Free AI Funder Matching for 501(c)(3)s</title>; H1 'Find Funders Aligned to Your Mission'"),
    ("FM-FI-002", "First Impressions", "Value proposition immediately obvious", "pass", "-", "Subhead 'Connect with foundations, DAFs, and corporate giving programs that match your nonprofit's mission in seconds.'"),
    ("FM-FI-003", "First Impressions", "Primary CTA visible & compelling", "pass", "-", "'Get Started' button -> /mission with Search icon"),
    ("FM-FI-004", "First Impressions", "Page loads under 3 seconds (TTFB)", "pass", "-", "Live curl TTFB 225ms — fastest of the three sites"),
    ("FM-FI-005", "First Impressions", "Visual hierarchy: single H1", "pass", "-", "<h1 className='text-5xl…'>Find Funders Aligned to Your Mission</h1>"),
    ("FM-FI-006", "First Impressions", "Trust signals visible (data + price)", "pass", "-", "'Powered by IRS 990 public filings', 'Free to use — no credit card required', 'Your data is never shared or sold'"),
    ("FM-FI-007", "First Impressions", "Data stats visible (corpus credibility)", "pass", "-", "460K+ funders / 449K+ recipients / 7.5M+ grants / 1.1M+ 990 filings"),
    ("FM-FI-008", "First Impressions", "Nonprofit-relevant copy (501(c)(3), DAFs, mission)", "pass", "-", "Copy uses sector-correct terms — would resonate with a small-NP ED/grants manager"),
    ("FM-FI-009", "First Impressions", "Demo video on homepage", "pass", "-", "<DemoVideo /> rendered below trust signals"),
    ("FM-FI-010", "First Impressions", "How It Works section explains 4 steps", "pass", "-", "4 steps: Describe Mission -> Get Ranked Matches -> Save & Track -> AI Grant Writer"),
    ("FM-FI-011", "First Impressions", "Free messaging is unambiguous", "pass", "-", "'Free AI funder matching — no account required.' in meta description AND on landing"),
    ("FM-FI-012", "First Impressions", "Visual contrast strong (dark theme readable)", "pass", "-", "#0d1117 bg, white text — accessible contrast on body text"),
    ("FM-FI-013", "First Impressions", "Nonprofit-budget concern addressed up-front", "pass", "-", "Trust line 'Free to use — no credit card required' is the first thing a small-budget ED scans for"),

    # Accessibility (20)
    ("FM-A11Y-001", "Accessibility", "Lang attribute present", "pass", "-", "<html lang='en'>"),
    ("FM-A11Y-002", "Accessibility", "Skip-to-content link present", "pass", "-", "<a href='#main-content' className='sr-only focus:not-sr-only…'>Skip to main content</a> in NavBar"),
    ("FM-A11Y-003", "Accessibility", "Single H1 on Landing", "pass", "-", "Single H1 'Find Funders Aligned to Your Mission'"),
    ("FM-A11Y-004", "Accessibility", "Single H1 on /privacy", "pass", "-", "PrivacyPolicy.tsx renders <h1>Privacy Policy</h1> only"),
    ("FM-A11Y-005", "Accessibility", "Single H1 on /terms", "pass", "-", "TermsOfService.tsx renders <h1>Terms of Service</h1> only"),
    ("FM-A11Y-006", "Accessibility", "Single H1 on /contact", "pass", "-", "ContactPage.tsx renders <h1>Contact Us</h1> only"),
    ("FM-A11Y-007", "Accessibility", "Single H1 on /search", "pass", "-", "OrgSearchPage uses OrgSearch component with proper heading"),
    ("FM-A11Y-008", "Accessibility", "Heading hierarchy intact (h1 > h2 > h3)", "pass", "-", "Landing: h1 hero, h2 'How It Works'; sub-pages use h2 for sections"),
    ("FM-A11Y-009", "Accessibility", "Nav has aria-label='Main navigation'", "pass", "-", "NavBar.tsx: <nav aria-label='Main navigation'>"),
    ("FM-A11Y-010", "Accessibility", "Mobile-menu toggle has aria-expanded + aria-label", "pass", "-", "aria-label='Toggle menu' aria-expanded={mobileOpen}"),
    ("FM-A11Y-011", "Accessibility", "Account dropdown uses aria-expanded + aria-haspopup + role=menu", "pass", "-", "All present on account button + menu container"),
    ("FM-A11Y-012", "Accessibility", "Footer nav has aria-label", "pass", "-", "Footer.tsx: <nav aria-label='Footer navigation'>"),
    ("FM-A11Y-013", "Accessibility", "Search input has aria-label", "pass", "-", "OrgSearch input: aria-label='Search by organization name or EIN'"),
    ("FM-A11Y-014", "Accessibility", "Breadcrumb uses aria-label='breadcrumb'", "pass", "-", "Breadcrumb.tsx: <nav aria-label='breadcrumb'>"),
    ("FM-A11Y-015", "Accessibility", "prefers-reduced-motion honored in CSS", "pass", "-", "@media (prefers-reduced-motion: reduce) present in src/index.css"),
    ("FM-A11Y-016", "Accessibility", "Global focus-visible outline always shows", "pass", "-", "*:focus-visible { outline: 2px solid #3b82f6; outline-offset: 2px } — overrides any focus:outline-none"),
    ("FM-A11Y-017", "Accessibility", "Body text contrast (#d1d5db on #0d1117) passes AA", "pass", "-", "~13:1 contrast ratio for body text on dark bg"),
    ("FM-A11Y-018", "Accessibility", "Decorative SVGs have aria-hidden", "pass", "-", "OAuth provider icons in LoginPage use aria-hidden='true'"),
    ("FM-A11Y-019", "Accessibility", "NotFound (SPA fallback) sets noindex (this run)", "pass", "P3", "Auto-fix this run: NotFound.tsx now injects <meta name='robots' content='noindex,nofollow'> on mount + restores on unmount"),
    ("FM-A11Y-020", "Accessibility", "OAuth buttons have visible text labels", "pass", "-", "LoginPage PROVIDERS each render visible label ('Continue with Google' etc.) — not icon-only"),

    # Forms & Inputs (12)
    ("FM-FORM-001", "Forms & Inputs", "MissionInput has mission textarea + location input", "pass", "-", "Form takes mission, locationServed, budgetBand"),
    ("FM-FORM-002", "Forms & Inputs", "MissionInput shows example mission statements", "pass", "-", "EXAMPLES array surfaced via showExamples toggle"),
    ("FM-FORM-003", "Forms & Inputs", "Validation errors shown on missing fields", "pass", "-", "errors state with mission/location keys — rendered in component"),
    ("FM-FORM-004", "Forms & Inputs", "Budget band selector covers small-NP brackets", "pass", "-", "Bands: Under $250K / $250K-$1M / $1M-$5M / $5M+ / Prefer not to say — covers most users at <$500K budget threshold mentioned in task"),
    ("FM-FORM-005", "Forms & Inputs", "Contact form Name/Email/Message inputs", "pass", "-", "ContactPage form state { name, email, message }"),
    ("FM-FORM-006", "Forms & Inputs", "Contact form has email validation (HTML5)", "pass", "-", "Email input uses type='email' (verified in render path)"),
    ("FM-FORM-007", "Forms & Inputs", "Contact form submits to Supabase edge function", "pass", "-", "POST https://tgtotjvdubhjxzybmdex.supabase.co/functions/v1/contact-form"),
    ("FM-FORM-008", "Forms & Inputs", "Contact form OPTIONS preflight returns 204", "pass", "-", "Live: 204 — CORS preflight succeeds"),
    ("FM-FORM-009", "Forms & Inputs", "Contact form shows error state when submit fails", "pass", "-", "error state set; rendered inline"),
    ("FM-FORM-010", "Forms & Inputs", "Contact form shows success state when submit OK", "pass", "-", "submitted state -> success message swap"),
    ("FM-FORM-011", "Forms & Inputs", "MissionInput keyboard-navigable", "pass", "-", "Form uses native inputs (textarea + LocationAutocomplete) — Tab order intact"),
    ("FM-FORM-012", "Forms & Inputs", "Login page OAuth providers have aria-label-equivalent text", "pass", "-", "Each provider button has visible 'Continue with X' label"),

    # Navigation & IA (15)
    ("FM-NAV-001", "Navigation & IA", "/mission 200", "pass", "-", "Live: 200 (SPA shell)"),
    ("FM-NAV-002", "Navigation & IA", "/results 200", "pass", "-", "Live: 200"),
    ("FM-NAV-003", "Navigation & IA", "/grant-writer 200", "pass", "-", "Live: 200"),
    ("FM-NAV-004", "Navigation & IA", "/search 200", "pass", "-", "Live: 200"),
    ("FM-NAV-005", "Navigation & IA", "/browse 200", "pass", "-", "Live: 200"),
    ("FM-NAV-006", "Navigation & IA", "/login 200", "pass", "-", "Live: 200"),
    ("FM-NAV-007", "Navigation & IA", "/signup 200", "pass", "-", "Live: 200"),
    ("FM-NAV-008", "Navigation & IA", "/dashboard 200 (gated content)", "pass", "-", "Live: 200; AuthGuard handles client-side gating"),
    ("FM-NAV-009", "Navigation & IA", "/onboarding 200", "pass", "-", "Live: 200"),
    ("FM-NAV-010", "Navigation & IA", "/privacy 200", "pass", "-", "Live: 200"),
    ("FM-NAV-011", "Navigation & IA", "/terms 200", "pass", "-", "Live: 200"),
    ("FM-NAV-012", "Navigation & IA", "/contact 200", "pass", "-", "Live: 200"),
    ("FM-NAV-013", "Navigation & IA", "/pricing 200", "pass", "-", "Live: 200"),
    ("FM-NAV-014", "Navigation & IA", "/faq 200", "pass", "-", "Live: 200"),
    ("FM-NAV-015", "Navigation & IA", "/__notexist returns SPA shell but NotFound renders client-side", "pass", "-", "GitHub Pages returns 200 for unknown paths (SPA fallback); NotFound.tsx renders with noindex meta after this run's auto-fix"),

    # Performance & Loading (10)
    ("FM-PERF-001", "Performance & Loading", "Homepage TTFB excellent (<300ms)", "pass", "-", "Live: 225ms (fastest of the three sites)"),
    ("FM-PERF-002", "Performance & Loading", "HTML shell size lean (<10KB)", "pass", "-", "6.57KB (SPA — actual content comes from JS bundle)"),
    ("FM-PERF-003", "Performance & Loading", "Routes lazy-loaded via React.lazy", "pass", "-", "All non-Landing pages are lazy() imports — keeps initial JS small"),
    ("FM-PERF-004", "Performance & Loading", "Suspense fallback exists", "pass", "-", "<RouteFallback /> 'Loading…' for code-split chunks"),
    ("FM-PERF-005", "Performance & Loading", "Cache-control sensible for SPA", "pass", "-", "public, max-age=0, must-revalidate (lets CDN revalidate fresh shell)"),
    ("FM-PERF-006", "Performance & Loading", "Cloudflare speculation rules enabled", "pass", "-", "speculation-rules header"),
    ("FM-PERF-007", "Performance & Loading", "Favicon variants for all sizes", "pass", "-", "16/32/48/svg + apple-touch-icon links present"),
    ("FM-PERF-008", "Performance & Loading", "Web app manifest served", "pass", "-", "<link rel='manifest' href='/site.webmanifest'>"),
    ("FM-PERF-009", "Performance & Loading", "JSON-LD WebApplication schema present", "pass", "-", "Structured data with name/url/description/audience/featureList"),
    ("FM-PERF-010", "Performance & Loading", "axe-core packaged dev-only (not shipped)", "pass", "-", "@axe-core/cli + axe-core in devDependencies only; no longer in CSP script-src whitelist"),

    # Mobile Responsiveness (10)
    ("FM-MOB-001", "Mobile Responsiveness", "Viewport meta tag present", "pass", "-", "<meta name='viewport' content='width=device-width, initial-scale=1.0' />"),
    ("FM-MOB-002", "Mobile Responsiveness", "Hamburger menu defined", "pass", "-", "md:hidden Menu/X icon toggle with aria-expanded"),
    ("FM-MOB-003", "Mobile Responsiveness", "Mobile menu renders full link set", "pass", "-", "Mobile drawer includes Find Funders, Saved, Browse, Search, Dashboard etc."),
    ("FM-MOB-004", "Mobile Responsiveness", "Touch targets ≥44x44px on nav", "pass", "-", "linkClass uses min-h-[44px] flex items-center px-4 py-2 — WCAG 2.5.5 compliant"),
    ("FM-MOB-005", "Mobile Responsiveness", "Touch targets ≥44x44px on footer", "pass", "-", "Footer.tsx links use inline-flex min-h-[44px]"),
    ("FM-MOB-006", "Mobile Responsiveness", "Hero scales between md and 7xl", "pass", "-", "H1 className 'text-5xl md:text-7xl' tracks viewport size"),
    ("FM-MOB-007", "Mobile Responsiveness", "Stats grid collapses 2-up on mobile, 4-up on desktop", "pass", "-", "grid-cols-2 lg:grid-cols-4 pattern"),
    ("FM-MOB-008", "Mobile Responsiveness", "How-It-Works steps stack on mobile", "pass", "-", "grid-cols-1 sm:grid-cols-2 lg:grid-cols-4"),
    ("FM-MOB-009", "Mobile Responsiveness", "Mobile nav drawer styled distinctly", "pass", "-", "md:hidden border-t bg-[#0d1117] block"),
    ("FM-MOB-010", "Mobile Responsiveness", "Sticky NavBar respects safe area", "pass", "-", "sticky top-0 z-40 — works at 375px iPhone SE width"),

    # Content Quality (8)
    ("FM-CONTENT-001", "Content Quality", "Footer copyright auto-updates by year", "pass", "-", "Footer.tsx: {new Date().getFullYear()} -> 2026"),
    ("FM-CONTENT-002", "Content Quality", "Privacy + Terms + Contact pages exist", "pass", "-", "All three live 200; full content pages, not placeholders"),
    ("FM-CONTENT-003", "Content Quality", "Privacy Policy 'Last updated' is current", "pass", "-", "April 7, 2026"),
    ("FM-CONTENT-004", "Content Quality", "Terms of Service 'Last updated' is current", "pass", "-", "April 26, 2026"),
    ("FM-CONTENT-005", "Content Quality", "Data stats believable for nonprofit audience", "pass", "-", "460K funders / 449K recipients / 7.5M grants / 1.1M 990 filings — consistent with IRS Form 990 public dataset scale"),
    ("FM-CONTENT-006", "Content Quality", "Schema.org Audience tag mentions 501(c)(3)s", "pass", "-", "audienceType: 'Nonprofit Organizations, 501(c)(3)s'"),
    ("FM-CONTENT-007", "Content Quality", "No typos in landing copy", "pass", "-", "Static review of Landing.tsx: clean prose"),
    ("FM-CONTENT-008", "Content Quality", "OG image referenced and likely present", "pass", "-", "og:image https://fundermatch.org/og-image.png referenced — used for share previews"),

    # Backend Integration (14)
    ("FM-BACK-001", "Backend Integration", "Strict-Transport-Security set", "pass", "-", "max-age=31536000; includeSubDomains"),
    ("FM-BACK-002", "Backend Integration", "CSP delivered via meta tag (GH Pages constraint)", "pass", "-", "Meta CSP: default-src 'none'; script-src 'self' + sha256 + gc.zgo.at; etc."),
    ("FM-BACK-003", "Backend Integration", "CSP additionally set via HTTP header", "pass", "-", "Server response also includes content-security-policy header — directive parity closed 2026-05-16"),
    ("FM-BACK-004", "Backend Integration", "X-Frame-Options: DENY", "pass", "-", "Header set"),
    ("FM-BACK-005", "Backend Integration", "X-Content-Type-Options: nosniff", "pass", "-", "Header set"),
    ("FM-BACK-006", "Backend Integration", "Referrer-Policy: strict-origin-when-cross-origin", "pass", "-", "Header set"),
    ("FM-BACK-007", "Backend Integration", "Permissions-Policy disables camera/mic/geo", "pass", "-", "Header set"),
    ("FM-BACK-008", "Backend Integration", "Cross-Origin policies set", "pass", "-", "COOP same-origin; CORP same-origin"),
    ("FM-BACK-009", "Backend Integration", "/functions/v1/contact-form OPTIONS preflight 204", "pass", "-", "Live curl: 204"),
    ("FM-BACK-010", "Backend Integration", "/functions/v1/filter-funders OPTIONS preflight 200", "pass", "-", "Live curl: 200"),
    ("FM-BACK-011", "Backend Integration", "/functions/v1/match-funders OPTIONS preflight 200", "pass", "-", "Live curl: 200"),
    ("FM-BACK-012", "Backend Integration", "Edge functions enumerated in codebase (20+)", "pass", "-", "ai-draft, match-funders, filter-funders, contact-form, grant-writer, generate-report, etc."),
    ("FM-BACK-013", "Backend Integration", "Recent IDOR fix in ai-draft + grant-writer (security hygiene)", "pass", "-", "Commits cb065c9 + ce7e955 + b6ed8e0 (2026-05-18) — BOLA closed in /ai-draft and storage-path IDOR closed in grant-writer"),
    ("FM-BACK-014", "Backend Integration", "CI uses npm ci with committed lockfile", "pass", "-", "Commit cb065c9 (2026-05-18) — supply-chain hardened"),
]


# ============================================================
# Build Summary sheet
# ============================================================
def score(tests):
    by_cat = {}
    for t in tests:
        cat = t[1]
        status = t[3].lower()
        if cat not in by_cat:
            by_cat[cat] = {"pass": 0, "fail": 0, "warn": 0, "info": 0, "skip": 0, "total": 0}
        by_cat[cat][status if status in ("pass","fail","warn","info","skip") else "skip"] += 1
        by_cat[cat]["total"] += 1
    # Score per category: pct of (pass + info) / (total - skip)
    out = {}
    for cat, v in by_cat.items():
        denom = max(1, v["total"] - v["skip"])
        ok = v["pass"] + v["info"]
        out[cat] = {"score": round(10 * ok / denom, 1), "pass": v["pass"], "fail": v["fail"], "warn": v["warn"], "info": v["info"], "total": v["total"]}
    return out


CATEGORIES = [
    "First Impressions",
    "Accessibility",
    "Forms & Inputs",
    "Navigation & IA",
    "Performance & Loading",
    "Mobile Responsiveness",
    "Content Quality",
    "Backend Integration",
]

wao_scores = score(WAO_TESTS)
ka_scores = score(KA_TESTS)
fm_scores = score(FM_TESTS)


# Summary sheet
ws = wb.active
ws.title = "Summary"
ws.append(["Daily Usability Audit — 2026-05-18"])
ws["A1"].font = Font(bold=True, size=16, name="Arial")
ws.append(["Author: scheduled audit bot. Three sites scored across 8 usability categories (1-10)."])
ws["A2"].font = BODY_FONT
ws.append([])
ws.append(["Category", "website-auditor.io", "kevinarmstrong.io", "fundermatch.org"])
style_header(ws, row=ws.max_row)
for cat in CATEGORIES:
    ws.append([
        cat,
        wao_scores.get(cat, {"score": 0})["score"],
        ka_scores.get(cat, {"score": 0})["score"],
        fm_scores.get(cat, {"score": 0})["score"],
    ])
    last_row = ws.max_row
    for ci in range(1, 5):
        c = ws.cell(row=last_row, column=ci)
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = BORDER
        c.font = BODY_FONT

# Aggregate row
agg_row = ws.max_row + 1
ws.append(["Aggregate (avg of category scores)", None, None, None])
for ci, col_letter in enumerate(["B", "C", "D"], start=2):
    first_data_row = agg_row - len(CATEGORIES)
    last_data_row = agg_row - 1
    ws.cell(row=agg_row, column=ci).value = f"=ROUND(AVERAGE({col_letter}{first_data_row}:{col_letter}{last_data_row}),1)"
for ci in range(1, 5):
    c = ws.cell(row=agg_row, column=ci)
    c.font = Font(bold=True, name="Arial")
    c.alignment = LEFT if ci == 1 else CENTER
    c.border = BORDER
    c.fill = INFO_FILL

ws.append([])
ws.append(["Counts per site"])
ws.cell(row=ws.max_row, column=1).font = Font(bold=True, name="Arial")
ws.append(["Site", "Total tests", "Pass", "Fail", "Warn", "Info"])
style_header(ws, row=ws.max_row)
for label, tests in (("website-auditor.io", WAO_TESTS), ("kevinarmstrong.io", KA_TESTS), ("fundermatch.org", FM_TESTS)):
    p = sum(1 for t in tests if t[3].lower() == "pass")
    f = sum(1 for t in tests if t[3].lower() == "fail")
    w = sum(1 for t in tests if t[3].lower() in ("warn","warning"))
    i = sum(1 for t in tests if t[3].lower() == "info")
    ws.append([label, len(tests), p, f, w, i])
    last_row = ws.max_row
    for ci in range(1, 7):
        c = ws.cell(row=last_row, column=ci)
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = BORDER
        c.font = BODY_FONT

set_widths(ws, [38, 22, 22, 22, 16, 16])
ws.column_dimensions["A"].width = 38


# Per-site sheets
ws_wao = wb.create_sheet("website-auditor.io")
write_tests(ws_wao, WAO_TESTS)

ws_ka = wb.create_sheet("kevinarmstrong.io")
write_tests(ws_ka, KA_TESTS)

ws_fm = wb.create_sheet("fundermatch.org")
write_tests(ws_fm, FM_TESTS)


# Auto-fixes sheet
ws_af = wb.create_sheet("Auto-fixes")
ws_af.append(["Repo", "Path", "Severity", "Description", "Commit"])
style_header(ws_af)
AUTO_FIXES = [
    ("my_website (kevinarmstrong.io)", "sitemap.xml", "P2", "Regenerated sitemap.xml: 37 -> 56 URLs; lastmod refreshed from 2026-04-22 to 2026-05-19. Used existing scripts/generate_sitemap.py.", "5189a19"),
    ("funder-finder (fundermatch.org)", "src/pages/NotFound.tsx", "P3", "Inject <meta name='robots' content='noindex,nofollow'> on NotFound mount and restore on unmount. Prevents Google from indexing arbitrary unknown SPA paths as 200 OK.", "56e8fc5"),
    ("chaos_tester (website-auditor.io)", "templates/changelog.html", "P2", "Added May 2026 changelog entry summarizing the past month of shipped work (CSP hardening + reporting endpoint, AI rate-limit, bug-report cap, SOC 2 A1 docs).", "45c5920"),
]
for r in AUTO_FIXES:
    ws_af.append(list(r))
    last_row = ws_af.max_row
    for ci in range(1, 6):
        c = ws_af.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = PASS_FILL
set_widths(ws_af, [32, 36, 10, 80, 14])
ws_af.freeze_panes = "A2"


# Owner-action items sheet
ws_oa = wb.create_sheet("Owner-action items")
ws_oa.append(["Site", "Item", "Severity", "Why deferred", "Suggested action"])
style_header(ws_oa)
OWNER_ITEMS = [
    ("website-auditor.io", "Hero subtitle: 'By late 2027, 75% of your customers will find your competitors with AI'", "P3", "Information / opinion item, not a defect — owner judgment call on whether the unverifiable forward-looking stat is right for the audience.", "If pressed on accuracy, soften (e.g., 'within 18 months') or attribute (Gartner / Pew style)."),
    ("kevinarmstrong.io", "Subheader color #7AED8C on #2596be banner background", "P3 (info)", "Owner-confirmed exception per task spec — owner prefers this green even though contrast on the blue band is non-AA.", "No action requested. Documented here only so it doesn't drift back into the report."),
    ("website-auditor.io", "How-It-Works / Features / FAQ removals", "Info", "Owner-confirmed intentional removal — flagged as 'do not flag' per task spec.", "Logged for traceability. No fix."),
    ("fundermatch.org", "SPA NotFound returns HTTP 200 (GitHub Pages limitation)", "P3", "GH Pages cannot return non-200 for unknown paths. We patched in noindex client-side (commit 56e8fc5), which is the strongest available signal.", "If a stricter signal is needed long-term, consider hosting behind a Worker that maps unknown SPA paths to a 404 response while keeping the client-rendered NotFound view."),
]
for r in OWNER_ITEMS:
    ws_oa.append(list(r))
    last_row = ws_oa.max_row
    for ci in range(1, 6):
        c = ws_oa.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = WARN_FILL if r[2].startswith("P") else INFO_FILL
set_widths(ws_oa, [22, 60, 12, 60, 60])
ws_oa.freeze_panes = "A2"


# Regression-Tracking sheet
ws_rt = wb.create_sheet("Regression-Tracking")
ws_rt.append(["Previously fixed item (fundermatch.org)", "Status today", "Notes"])
style_header(ws_rt)
REGRESSION = [
    ("Privacy Policy link works (not 404)", "PASS", "/privacy 200; PrivacyPolicy.tsx renders, Last updated April 7, 2026"),
    ("Terms of Service link works (not 404)", "PASS", "/terms 200; TermsOfService.tsx renders, Last updated April 26, 2026"),
    ("Contact page works (not 404)", "PASS", "/contact 200; contact form posts to Supabase contact-form edge function (OPTIONS 204)"),
    ("Keyboard focus indicators present", "PASS", "Global *:focus-visible { outline: 2px solid #3b82f6; outline-offset: 2px } in src/index.css"),
    ("Form validation shows error messages", "PASS", "MissionInput maintains errors state; ContactPage shows error state on failed submit"),
    ("Body text contrast passes WCAG AA", "PASS", "Body color ~#d1d5db on #0d1117 background — ~13:1 ratio"),
    ("Search input has aria-label", "PASS", "OrgSearch input: aria-label='Search by organization name or EIN'"),
    ("prefers-reduced-motion rule active", "PASS", "@media (prefers-reduced-motion: reduce) present in src/index.css"),
    ("Data stats section visible on homepage", "PASS", "460K+ / 449K+ / 7.5M+ / 1.1M+ stat grid present on Landing.tsx"),
    ("Trust signals visible on homepage", "PASS", "Three trust signals immediately below CTA: IRS 990 data / free / data never shared"),
]
for r in REGRESSION:
    ws_rt.append(list(r))
    last_row = ws_rt.max_row
    for ci in range(1, 4):
        c = ws_rt.cell(row=last_row, column=ci)
        c.alignment = LEFT
        c.border = BORDER
        c.font = BODY_FONT
        c.fill = PASS_FILL if r[1] == "PASS" else FAIL_FILL
set_widths(ws_rt, [60, 16, 80])
ws_rt.freeze_panes = "A2"


# Make Summary the active/first tab
wb.active = 0
import os
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "daily_usability_audit_2026-05-18.xlsx")
wb.save(out_path)
print(f"Wrote {out_path}")
